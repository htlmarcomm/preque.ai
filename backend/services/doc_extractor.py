import os
from typing import List, Dict
import pypdf
import docx
import openpyxl
import logging
import io

logger = logging.getLogger(__name__)

def extract_text(filepath: str) -> List[Dict[str, str]]:
    """
    Extracts text from a file, returning a list of dicts:
    [{"page_or_sheet": str, "text": str}, ...]
    Supported extensions: .pdf, .xlsx, .xls, .docx
    """
    if not os.path.exists(filepath):
        logger.warning(f"File not found: {filepath}")
        return []

    ext = os.path.splitext(filepath)[1].lower()
    
    if ext == ".pdf":
        return _extract_pdf(filepath)
    elif ext in [".xlsx", ".xls"]:
        return _extract_excel(filepath)
    elif ext == ".docx":
        return _extract_docx(filepath)
    else:
        logger.warning(f"Unsupported file extension for extraction: {ext}")
        return []

def _extract_pdf(filepath: str) -> List[Dict[str, str]]:
    results = []
    try:
        with open(filepath, "rb") as f:
            reader = pypdf.PdfReader(f)
            for i, page in enumerate(reader.pages):
                text = page.extract_text()
                if text and text.strip():
                    results.append({"page_or_sheet": f"Page {i + 1}", "text": text.strip()})
    except Exception as e:
        logger.warning(f"Failed to extract text from PDF {filepath}: {e}")
        
    if not results:
        logger.info(f"Native extraction failed for {filepath}, falling back to Vision API.")
        results = _extract_pdf_vision(filepath)
        
    return results

def _extract_pdf_vision(filepath: str) -> List[Dict[str, str]]:
    import tempfile, shutil, base64
    from pdf2image import convert_from_path
    from routers.forms import openai_client
    
    results = []
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".pdf")
    os.close(tmp_fd)
    try:
        shutil.copy(filepath, tmp_path)
        pages = convert_from_path(tmp_path, dpi=150)
        
        for i, page in enumerate(pages):
            img_byte_arr = io.BytesIO()
            page.save(img_byte_arr, format='PNG')
            b64 = base64.standard_b64encode(img_byte_arr.getvalue()).decode('utf-8')
            
            prompt = (
                "You are a transcription assistant. Transcribe the text from this document image exactly as it appears. "
                "Do not add any additional commentary, just output the raw text."
            )
            
            try:
                resp = openai_client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[
                        {"role": "system", "content": prompt},
                        {
                            "role": "user", 
                            "content": [{"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}", "detail": "high"}}]
                        }
                    ],
                    max_tokens=4000
                )
                text = resp.choices[0].message.content.strip()
                if text:
                    results.append({"page_or_sheet": f"Page {i + 1}", "text": text})
            except Exception as e:
                logger.warning(f"Vision transcription failed for page {i+1} of {filepath}: {e}")
                
    except Exception as e:
        logger.warning(f"Vision extraction pipeline failed for {filepath}: {e}")
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
            
    return results

def _extract_excel(filepath: str) -> List[Dict[str, str]]:
    results = []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_text = []
            for row in ws.iter_rows(values_only=True):
                cleaned = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if cleaned:
                    rows_text.append(" | ".join(cleaned))
            if rows_text:
                results.append({"page_or_sheet": sheet_name, "text": "\n".join(rows_text)})
        wb.close()
    except Exception as e:
        logger.warning(f"Failed to extract text from Excel {filepath}: {e}")
    return results

def _extract_docx(filepath: str) -> List[Dict[str, str]]:
    results = []
    try:
        doc = docx.Document(filepath)
        paras = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
        if paras:
            results.append({"page_or_sheet": "Document", "text": "\n\n".join(paras)})
    except Exception as e:
        logger.warning(f"Failed to extract text from DOCX {filepath}: {e}")
    return results

def chunk_text(pages: List[Dict[str, str]], parent_chunk_size: int = 1000, parent_overlap: int = 200, child_chunk_size: int = 250, child_overlap: int = 50) -> List[Dict]:
    """
    Splits text from pages into overlapping parent chunks, and then splits those into child chunks.
    Returns: [{"sheet_or_page": str, "chunk_index": int, "text": str, "parent_text": str}, ...]
    """
    chunks = []
    global_chunk_idx = 0
    
    for page in pages:
        text = page["text"]
        page_name = page["page_or_sheet"]
        
        p_start = 0
        text_length = len(text)
        
        while p_start < text_length:
            p_end = p_start + parent_chunk_size
            
            # If not at the end, try to break at a newline or space for parent
            if p_end < text_length:
                last_newline = text.rfind('\n', p_start, p_end)
                if last_newline != -1 and (p_end - last_newline) < 200:
                    p_end = last_newline + 1
                else:
                    last_space = text.rfind(' ', p_start, p_end)
                    if last_space != -1 and (p_end - last_space) < 100:
                        p_end = last_space + 1

            parent_str = text[p_start:p_end].strip()
            
            if parent_str:
                # Now split parent_str into child chunks
                c_start = 0
                parent_length = len(parent_str)
                
                while c_start < parent_length:
                    c_end = c_start + child_chunk_size
                    
                    if c_end < parent_length:
                        last_c_newline = parent_str.rfind('\n', c_start, c_end)
                        if last_c_newline != -1 and (c_end - last_c_newline) < 50:
                            c_end = last_c_newline + 1
                        else:
                            last_c_space = parent_str.rfind(' ', c_start, c_end)
                            if last_c_space != -1 and (c_end - last_c_space) < 50:
                                c_end = last_c_space + 1
                                
                    child_str = parent_str[c_start:c_end].strip()
                    if child_str:
                        chunks.append({
                            "sheet_or_page": page_name,
                            "chunk_index": global_chunk_idx,
                            "text": child_str,
                            "parent_text": parent_str
                        })
                        global_chunk_idx += 1
                        
                    c_start = c_end - child_overlap
                    if c_start <= 0 or (c_end >= parent_length):
                        c_start = c_end
            
            p_start = p_end - parent_overlap
            if p_start <= 0 or (p_end >= text_length):
                p_start = p_end
            
    return chunks
