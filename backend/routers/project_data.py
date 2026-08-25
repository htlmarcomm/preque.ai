import io
import re
import os
import hashlib
from datetime import datetime
from fastapi import APIRouter, Depends, UploadFile, File, Query, HTTPException, Request
from sqlalchemy.orm import Session
from sqlalchemy import func

from models.database import get_db, ProjectDataRecord, ProjectDataColumn, ProjectDataSheet
from utils import sanitize_filename
import openpyxl

router = APIRouter()

def slugify(text):
    text = str(text).lower()
    text = re.sub(r'[^a-z0-9]+', '-', text)
    return text.strip('-')

def normalize_column_key(text):
    return slugify(text).replace("-", "_")

def detect_header_row(ws, max_scan_rows=10):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_scan_rows:
            break
        str_count = sum(1 for cell in row if cell is not None and str(cell).strip() != "" and isinstance(cell, str))
        non_empty = sum(1 for cell in row if cell is not None and str(cell).strip() != "")
        if non_empty > 0 and (str_count >= 3 or str_count / non_empty >= 0.5):
            return i
    return 0

@router.post("/import")
async def import_project_data(file: UploadFile = File(...), db: Session = Depends(get_db)):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    
    upload_dir = "uploads/project_data_imports"
    os.makedirs(upload_dir, exist_ok=True)
    timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    safe_filename = f"{timestamp}_{sanitize_filename(file.filename)}"
    file_path = os.path.join(upload_dir, safe_filename)
    with open(file_path, "wb") as f:
        f.write(contents)
    
    source_file = file.filename
    sheets_processed = 0
    rows_ingested = 0
    rows_skipped_duplicate = 0
    new_columns_learned = []
    
    existing_columns = {col.column_key: col for col in db.query(ProjectDataColumn).all()}
    columns_to_update = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if ws.max_row == 0 or ws.max_column == 0:
            continue
            
        header_row_idx = detect_header_row(ws)
        
        headers = []
        column_order = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < header_row_idx:
                continue
                
            if i == header_row_idx:
                sheets_processed += 1
                seen_keys = set()
                for h in row:
                    if h is not None and str(h).strip() != "":
                        base_key = normalize_column_key(h)
                        key = base_key
                        counter = 2
                        while key in seen_keys:
                            key = f"{base_key}_{counter}"
                            counter += 1
                        seen_keys.add(key)
                        headers.append((key, str(h).strip()))
                        column_order.append({"key": key, "display_label": str(h).strip()})
                    else:
                        headers.append((None, None))
                        
                # Create or fetch ProjectDataSheet
                sheet_record = db.query(ProjectDataSheet).filter(
                    ProjectDataSheet.source_file == source_file,
                    ProjectDataSheet.source_sheet == sheet_name
                ).first()
                if not sheet_record:
                    sheet_record = ProjectDataSheet(
                        source_file=source_file,
                        source_sheet=sheet_name,
                        column_order=column_order,
                        parser_used="tabular",
                        row_count=0
                    )
                    db.add(sheet_record)
                    db.commit()
                    db.refresh(sheet_record)
                continue
            
            # Process data row
            if not any(cell is not None and str(cell).strip() != "" for cell in row):
                continue
                
            row_data = {}
            for col_idx, cell in enumerate(row):
                if col_idx >= len(headers):
                    break
                col_key, original_label = headers[col_idx]
                if col_key is None:
                    continue
                
                cell_val = str(cell).strip() if cell is not None else ""
                row_data[col_key] = cell_val
                
                if col_key not in existing_columns:
                    new_col = ProjectDataColumn(
                        column_key=col_key,
                        display_label=original_label,
                        first_seen_file=source_file,
                        times_seen=0
                    )
                    db.add(new_col)
                    existing_columns[col_key] = new_col
                    new_columns_learned.append(col_key)
                
                if col_key not in columns_to_update:
                    columns_to_update[col_key] = 0
                columns_to_update[col_key] += 1

            if not row_data:
                continue
                
            search_text = " ".join(str(v).lower() for v in row_data.values() if v)
            sorted_items = sorted(row_data.items())
            hash_input = f"{source_file}|{sheet_name}|{i+1}|{sorted_items}"
            row_hash = hashlib.sha256(hash_input.encode('utf-8')).hexdigest()
            
            existing = db.query(ProjectDataRecord).filter(ProjectDataRecord.row_hash == row_hash).first()
            if existing:
                rows_skipped_duplicate += 1
                continue
                
            record = ProjectDataRecord(
                source_file=source_file,
                source_sheet=sheet_name,
                row_number=i+1,
                primary_label=None, # Removed the aggressive primary label logic as requested
                data=row_data,
                search_text=search_text,
                row_hash=row_hash
            )
            db.add(record)
            rows_ingested += 1

        db.commit()
        
        # Update row count for this sheet
        sheet_record = db.query(ProjectDataSheet).filter(
            ProjectDataSheet.source_file == source_file,
            ProjectDataSheet.source_sheet == sheet_name
        ).first()
        if sheet_record:
            sheet_record.row_count = db.query(ProjectDataRecord).filter(
                ProjectDataRecord.source_file == source_file,
                ProjectDataRecord.source_sheet == sheet_name
            ).count()
            db.commit()

    for col_key, increment in columns_to_update.items():
        col = existing_columns[col_key]
        col.times_seen += increment
        
    db.commit()
    
    total_cols = db.query(ProjectDataColumn).count()
    return {
        "file": source_file,
        "sheets_processed": sheets_processed,
        "rows_ingested": rows_ingested,
        "rows_skipped_duplicate": rows_skipped_duplicate,
        "new_columns_learned": list(set(new_columns_learned)),
        "total_columns_now": total_cols
    }

@router.get("/files")
def get_files(db: Session = Depends(get_db)):
    sheets = db.query(ProjectDataSheet).order_by(ProjectDataSheet.first_uploaded_at.desc()).all()
    
    file_map = {}
    for sheet in sheets:
        if sheet.source_file not in file_map:
            file_map[sheet.source_file] = {
                "source_file": sheet.source_file,
                "first_uploaded_at": sheet.first_uploaded_at,
                "total_rows": 0,
                "sheets": []
            }
        file_map[sheet.source_file]["total_rows"] += sheet.row_count
        file_map[sheet.source_file]["sheets"].append({
            "source_sheet": sheet.source_sheet,
            "row_count": sheet.row_count,
            "parser_used": sheet.parser_used,
            "columns": sheet.column_order
        })
        
    return list(file_map.values())

@router.get("/files/{source_file}/sheets/{source_sheet}/records")
def get_sheet_records(
    source_file: str,
    source_sheet: str,
    request: Request,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    search: str = None,
    db: Session = Depends(get_db)
):
    query = db.query(ProjectDataRecord).filter(
        ProjectDataRecord.source_file == source_file,
        ProjectDataRecord.source_sheet == source_sheet
    )
    
    if search:
        query = query.filter(ProjectDataRecord.search_text.ilike(f"%{search}%"))
        
    for key, val in request.query_params.items():
        if key.startswith("filter_"):
            col_key = key[7:]
            # JSON extract SQLite specific format: JSON_EXTRACT(data, '$."key"')
            query = query.filter(func.json_extract(ProjectDataRecord.data, f'$."{col_key}"') == val)
            
    total = query.count()
    records = query.order_by(ProjectDataRecord.row_number).offset((page - 1) * page_size).limit(page_size).all()
    
    sheet = db.query(ProjectDataSheet).filter(
        ProjectDataSheet.source_file == source_file,
        ProjectDataSheet.source_sheet == source_sheet
    ).first()
    columns = sheet.column_order if sheet else []
    
    return {
        "records": [
            {
                "id": r.id,
                "row_number": r.row_number,
                "data": r.data
            } for r in records
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
        "columns": columns
    }

@router.get("/files/{source_file}/sheets/{source_sheet}/filters")
def get_sheet_filters(
    source_file: str,
    source_sheet: str,
    db: Session = Depends(get_db)
):
    sheet = db.query(ProjectDataSheet).filter(
        ProjectDataSheet.source_file == source_file,
        ProjectDataSheet.source_sheet == source_sheet
    ).first()
    
    if not sheet or not sheet.column_order:
        return {"filters": []}
        
    records = db.query(ProjectDataRecord).filter(
        ProjectDataRecord.source_file == source_file,
        ProjectDataRecord.source_sheet == source_sheet
    ).all()
    
    filters = []
    for col in sheet.column_order:
        col_key = col["key"]
        distinct_vals = set()
        for r in records:
            val = r.data.get(col_key)
            if val and str(val).strip():
                distinct_vals.add(str(val).strip())
        
        if 2 <= len(distinct_vals) <= 20:
            filters.append({
                "key": col_key,
                "display_label": col["display_label"],
                "values": sorted(list(distinct_vals))
            })
            
    return {"filters": filters}

@router.get("/columns")
def get_columns(db: Session = Depends(get_db)):
    columns = db.query(ProjectDataColumn).order_by(ProjectDataColumn.times_seen.desc()).all()
    return [
        {
            "column_key": c.column_key,
            "display_label": c.display_label,
            "times_seen": c.times_seen,
            "first_seen_file": c.first_seen_file,
            "first_seen_at": c.first_seen_at
        } for c in columns
    ]

@router.get("/summary")
def get_summary(db: Session = Depends(get_db)):
    total_records = db.query(ProjectDataRecord).count()
    total_columns_learned = db.query(ProjectDataColumn).count()
    total_files = db.query(ProjectDataRecord.source_file).distinct().count()
    
    file_stats = db.query(
        ProjectDataRecord.source_file,
        func.count(ProjectDataRecord.id).label('record_count'),
        func.min(ProjectDataRecord.uploaded_at).label('first_uploaded_at')
    ).group_by(ProjectDataRecord.source_file).all()
    
    return {
        "total_records": total_records,
        "total_files": total_files,
        "total_columns_learned": total_columns_learned,
        "files": [
            {
                "source_file": f.source_file,
                "record_count": f.record_count,
                "first_uploaded_at": f.first_uploaded_at
            } for f in file_stats
        ]
    }

@router.delete("/records/{record_id}")
def delete_record(record_id: int, db: Session = Depends(get_db)):
    record = db.query(ProjectDataRecord).filter(ProjectDataRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
        
    db.delete(record)
    db.commit()
    return {"status": "deleted"}
