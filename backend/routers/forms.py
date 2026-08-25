from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from models.database import get_db, FilledForm
from openai import OpenAI
import openpyxl, io, os, json, re, base64, tempfile, subprocess, shutil
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from routers.project_files import read_excel_preview

router = APIRouter()
UPLOAD_DIR = "uploads/forms"
OUTPUT_DIR = "uploads/outputs"
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

openai_client = OpenAI(
    api_key=os.getenv("OPENAI_API_KEY"),
    # FIX (live bug -- every single AI call was 404ing): GitHub Models (the
    # original free provider this app was built on) was fully retired by
    # GitHub on 2026-07-30 -- not just moved, permanently shut down. Switched
    # to Google's Gemini API via its OpenAI-compatible endpoint instead: same
    # `openai` SDK calling convention (so none of the prompt/parsing code
    # needed to change), a genuinely free tier (no card required), and a much
    # higher free quota (1,500 req/day vs GitHub Models' old 50/day).
    base_url="https://generativelanguage.googleapis.com/v1beta/openai/",
    timeout=120.0,  # 2 min per API call
)
VISION_MODEL = "gemini-3.5-flash"
MINI_MODEL = "gemini-3.5-flash-lite"  # cheaper/faster text-only model for classification-style calls

# â”€â”€ Prompts â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

FILL_SYSTEM_PROMPT = """You are an expert form filler working for HTL Aircon Pvt Ltd, a MEP contractor.

You are filling ONE SPECIFIC SHEET from a multi-sheet Excel Pre-Qualification form. Stay focused only on this sheet — do not invent data for sheets you cannot see.

You will receive:
1. An image of this one Excel sheet (exactly as a human would see it)
2. A cell map for THIS SHEET ONLY — every cell shown as [COORD]='value', [COORD]=EMPTY, or [COORD]='--- Section: <Name> ---'
3. HTL Aircon's company data (Structured as a JSON object grouped by categories like "General", "Financials", etc.)

Your task — act like a human filling this one tab:
- Look at the image and the '--- Section: ... ---' markers in the cell map to understand which section you are in (e.g. Financial Info, General Info).
- Use this section context to navigate the structured JSON COMPANY DATA. For example, if you are in a "Financial" section, prioritize looking inside the "Financials" category of the JSON and match the correct Fiscal Year.
- If a field explicitly asks you to "attach", "enclose", or "provide a copy of" a certificate, license, or document (e.g. 'Attach PAN', 'Enclose GST', 'Upload ISO'), output the exact string "Attached". Do NOT output the actual registration number unless the cell specifically asks for the number itself.
- Use the cell map to find the EXACT coordinates of each empty answer cell ON THIS SHEET
- Match each empty answer cell to what it's asking using visual context. Pay close attention to headings followed by sub-headings/sub-rows — each sub-row is usually a distinct field that needs its own value (e.g. a "Annual Turnover" heading with sub-rows "FY 2022-23", "FY 2023-24", "FY 2024-25" each need their own figure; a "Contact Details" heading with sub-rows "Name", "Designation", "Phone" each need a different value).
- Fill each matched cell with the correct value from company data

Return ONLY a JSON object: {"D6": "HTL Aircon Pvt Ltd", "D8": "1996", ...}

Rules:
- Cell addresses must be exact (e.g. "D6") — get them from the cell map, not the image
- Be smart about matching: "Name of the firm" = Company Name, "GST No" = GSTIN, etc.
- For sub-rows under a heading, fill each sub-row's own adjacent cell — never repeat the same value into every sub-row unless they're genuinely asking for the same thing
- Cells belonging to a repeating list of DIFFERENT past projects or DIFFERENT past clients (e.g. "Client Reference -1", "-2", "-3" or a numbered project history table) are handled by a separate picker UI, not by you. If the cell map marks such rows as "(Reserved for project/reference picker)" leave them alone — do not guess a value for them, and never repeat one single company fact (like our own contact person) across what are clearly slots for several different external clients/projects.
- Skip cells where you have no matching data for THIS sheet — do not guess or fabricate
- Do NOT fill section headers, serial numbers, or label cells
- Do NOT fill cells that are clearly informational text, instructions, or declarations with no blank to answer
- Return ONLY the JSON, no markdown, no explanation"""

# FIX (P1 -- 50 req/day quota risk): this used to be three separate GPT-4o-mini
# calls per sheet (classify, column-policy, table-detect). GitHub Models' free
# tier caps at 50 requests/day, so a single multi-sheet form (e.g. 13 tabs)
# could burn the *entire* daily quota on classification alone, before a single
# vision-fill call ever ran. Combining all three questions into one call per
# sheet cuts that to a flat 1 call/sheet regardless of how many of the three
# questions are actually relevant.
SHEET_ANALYSIS_PROMPT = """You are analyzing ONE sheet/tab from a multi-sheet Excel Pre-Qualification form for HTL Aircon Pvt Ltd, a MEP contractor.

The sheet name is: "{sheet_name}"

Below is the cell map for JUST this sheet:
{cell_map}

Answer THREE questions about this sheet and return a single JSON object.

1. "classification": either
   - "FILLABLE" — it has labelled fields/questions (with or without sub-headings) that need
     company data filled into adjacent/nearby empty cells. This includes a heading followed
     by sub-fields (e.g. "Annual Turnover" as heading, then "2022-23", "2023-24" as sub-rows
     each needing a value).
   - "INFO_ONLY" — pure instructional text, a cover letter, declaration, index/table of
     contents, terms & conditions, or auto-calculated/summary sheet with no genuine blank
     answer cells belonging to HTL.

2. "columns": some forms reserve specific columns for the client's internal use only (e.g.
   "remarks by evaluator", "docs verified", "internal use only") versus columns meant for us
   (the vendor) to fill. Look at the header rows (usually row 1 to 5). If there's no clear
   distinction, or all columns seem fillable, return empty lists for both:
   {{"fillable_columns": ["B", "D"], "reserved_columns": ["E", "F", "G"]}}

3. "project_table": does this sheet contain a repeating table where each row represents ONE
   PROJECT (a list of client engagements/jobs)? Two distinct kinds:
   - TYPE "project_reference": each row asks for CLIENT CONTACT/REFERENCE details tied to a
     past project — client name, contact person, designation, phone, email — used so a
     reviewer can call a past client for a reference.
   - TYPE "project_details": each row asks for PROJECT-level facts — project name, location,
     client name, area, contract value/amount, start/completion dates, scope of work. If the
     table asks for project value, scope, duration, or area, classify as "project_details"
     even if it ALSO asks for a client contact person.
   The form might use different terminology instead of "Projects" or "References" — look for
   headers or preceding text like "Major work done", "Ongoing projects", "Projects completed",
   "Past performance", "Experience Record". If it asks for "Project Cost", "Duration of
   Project", "Scope of Work", "Location", "Client Name", treat it as a project table
   regardless of the exact heading. Look for consecutive numbered cells in a single column
   (e.g. [C7]='1', [C8]='2') indicating table rows, or clear repeating column headers. Also
   scan nearby text for an explicit max row count instruction (e.g. "list only 10 projects at
   max", "top 5 projects", "maximum 3 references") — extract that number if present, else null.
   If not a project table, just return {{"is_project_table": false}}.

Return ONLY this JSON object, no markdown, no explanation:
{{
  "classification": "FILLABLE",
  "columns": {{"fillable_columns": [], "reserved_columns": []}},
  "project_table": {{
    "is_project_table": true,
    "table_type": "project_reference",
    "subheading": "The exact heading text found (e.g. 'Major work done'), or null",
    "start_row": 7,
    "max_rows": 10,
    "mapping": {{
      "project_name": "D", "client_name": "E", "location": "F", "area_sqft": "G",
      "amount": "H", "start_date": "I", "completion_date": "J", "contact_name": "K",
      "contact_designation": "L", "contact_phone": "M", "contact_email": "N",
      "scope_of_work": "O"
    }}
  }}
}}
Only include mapping keys for columns that actually exist as headers in this sheet — do not
invent columns. Map any column asking for "Scope of Work" to the "scope_of_work" key.

CELL MAP:
{cell_map}
"""


def _clean(val) -> str:
    if val is None: return ""
    s = str(val).strip().replace('\n', ' ').replace('\t', ' ')
    return re.sub(r'\s+', ' ', s)[:100]


def _merge_continuation_cells(ws) -> set:
    """
    Returns the set of (row, col) tuples that are NON-ANCHOR continuations of a merged
    cell range (e.g. for a B8:C8 merge, this returns {(8, 3)} for C8, but not B8).

    FIX (P0 -- silent label corruption): these cells must never be treated as a real,
    independent, fillable cell. openpyxl reports their .value as always None, which
    previously made build_sheet_cell_map mark them "[COORD]=EMPTY" -- a completely
    reasonable-looking fill target to both Pass 1 (fuzzy alias match) and GPT-4o
    vision. But write_filled_excel_multi's merge-handling logic redirects any write
    aimed at one of these cells to the merge's TOP-LEFT ANCHOR -- which is exactly
    where the field's own label text lives. The result: filling what looks like a
    normal blank answer cell silently overwrites and destroys the form's own label
    (e.g. writing "1996" into what looked like an empty cell next to "Year of
    Establishment" actually replaced the words "Year of Establishment" with "1996").
    Excluding these coordinates from the cell map entirely means neither Pass 1 nor
    vision will ever propose filling them in the first place.
    """
    continuations = set()
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if (r, c) != (min_row, min_col):
                    continuations.add((r, c))
    return continuations


def _print_area_bounds(ws):
    """
    Returns (min_row, min_col, max_row, max_col) for the sheet's own defined
    print area, or None if it has none / it can't be parsed.

    FIX (P0 -- duplicate answers written into cells that were never meant to
    hold one): a form template can have stray extra blank columns sitting
    right next to the real answer column (e.g. a leftover draft column, or
    padding) that the template's own author explicitly excluded from
    print_area when they set it up. Pass 1 (deterministic matching) usually
    lands correctly in the real answer column since it's the FIRST empty
    cell found scanning outward, but GPT-4o vision looks at the rendered
    image and cell map together and, seeing another nearby EMPTY cell for
    the same visible field, would reasonably (if wrongly) fill that one too
    -- producing two slightly different values for what's really one
    answer. The template's own print_area is a strong, unambiguous signal
    for "this is the real submitted content" that was already sitting in
    the file, unused. Cells outside it are excluded from being offered as
    fillable at all.
    """
    from openpyxl.utils.cell import range_boundaries
    pa = getattr(ws, "print_area", None)
    if not pa:
        return None
    try:
        first_range = str(pa).split(',')[0].strip()
        if '!' in first_range:
            first_range = first_range.split('!', 1)[1]
        first_range = first_range.replace('$', '')
        min_col, min_row, max_col, max_row = range_boundaries(first_range)
        if None in (min_col, min_row, max_col, max_row):
            return None
        return (min_row, min_col, max_row, max_col)
    except Exception:
        return None


def _serial_number_columns(filled_cells: dict) -> set:
    """
    Detects columns that are pure sequential serial-number/index columns
    (e.g. column A holding 1, 2, 3, 4, 5... to mark numbered sections) so
    their EMPTY cells never get offered up as fillable answer slots.

    FIX (P0 -- real data written into an index column): the row-context
    matching used for Pass 1 doesn't care which column an empty cell sits
    in -- it just sees "this row's text says Mobile No." and fuzzy-matches
    that against any empty cell in the row, including the index/serial
    column itself if that happens to be blank on this particular sub-row.
    That silently overwrote what should stay a plain row number with real
    company data (e.g. a phone number landing in the "Sr. No." column).
    """
    by_col: dict = {}
    for (r, c), val in filled_cells.items():
        by_col.setdefault(c, []).append((r, val))

    serial_cols = set()
    for c, entries in by_col.items():
        if len(entries) < 3:
            continue
        entries.sort()
        nums = []
        ok = True
        for r, val in entries:
            if re.fullmatch(r'\d{1,3}', val.strip()):
                nums.append(int(val))
            else:
                ok = False
                break
        if ok and nums == sorted(nums) and len(set(nums)) == len(nums) and (nums[-1] - nums[0]) < len(nums) * 3:
            serial_cols.add(c)
    return serial_cols


def build_sheet_cell_map(ws, ws_formulas=None) -> tuple[str, set[str]]:
    """Build a compact cell map for a single worksheet, only including cells with values and adjacent empty cells.
    Returns (cell_map_str, protected_cells_set)."""
    filled_cells = {}
    protected_cells = set()
    real_max_r = 0
    real_max_c = 0
    merge_continuations = _merge_continuation_cells(ws)
    print_bounds = _print_area_bounds(ws)

    if ws_formulas:
        for v_row, f_row in zip(ws.iter_rows(), ws_formulas.iter_rows()):
            for v_cell, f_cell in zip(v_row, f_row):
                if f_cell.data_type == 'f':
                    protected_cells.add(v_cell.coordinate)
                if v_cell.value is not None:
                    real_max_r = max(real_max_r, v_cell.row)
                    real_max_c = max(real_max_c, v_cell.column)
                    filled_cells[(v_cell.row, v_cell.column)] = _clean(v_cell.value)
    else:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    real_max_r = max(real_max_r, cell.row)
                    real_max_c = max(real_max_c, cell.column)
                    filled_cells[(cell.row, cell.column)] = _clean(cell.value)
                
    if not filled_cells:
        return "", protected_cells
        
    if ws.max_row > real_max_r or ws.max_column > real_max_c:
        print(f"[TRIM] '{ws.title}': reported {ws.max_row}x{ws.max_column}, using real {real_max_r}x{real_max_c}")
        
    max_r = min(ws.max_row, real_max_r + 2)
    max_c = min(ws.max_column, real_max_c + 2)

    serial_cols = _serial_number_columns(filled_cells)

    # Detect horizontal merged cells as potential section headers
    section_headers = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if max_col > min_col and max_row == min_row: # Horizontal merge
            if (min_row, min_col) in filled_cells:
                text_val = filled_cells[(min_row, min_col)]
                if len(text_val) < 150:
                    section_headers[(min_row, min_col)] = text_val

    relevant_empty = set()
    for (r, c) in filled_cells.keys():
        for dr in range(-2, 3):
            for dc in range(-2, 3):
                nr, nc = r + dr, c + dc
                if 1 <= nr <= max_r and 1 <= nc <= max_c:
                    # FIX: never surface a merge-continuation cell, a cell
                    # outside the sheet's own print area, or a cell in a
                    # detected serial-number column as "EMPTY" -- see the
                    # docstrings on _merge_continuation_cells,
                    # _print_area_bounds, and _serial_number_columns above.
                    if (nr, nc) in filled_cells or (nr, nc) in merge_continuations:
                        continue
                    if print_bounds and not (print_bounds[0] <= nr <= print_bounds[2] and print_bounds[1] <= nc <= print_bounds[3]):
                        continue
                    if nc in serial_cols:
                        continue
                    relevant_empty.add((nr, nc))
                        
    rows_to_output = set(r for r, c in filled_cells.keys()) | set(r for r, c in relevant_empty)
    
    lines = []
    for r in sorted(list(rows_to_output)):
        cols_in_row = set(c for r_f, c in filled_cells.keys() if r_f == r) | \
                      set(c for r_e, c in relevant_empty if r_e == r)
        if not cols_in_row:
            continue
            
        parts = []
        for c in sorted(list(cols_in_row)):
            cell = ws.cell(row=r, column=c)
            coord = cell.coordinate
            if (r, c) in section_headers:
                parts.append(f"[{coord}]='--- Section: {section_headers[(r, c)]} ---'")
            elif (r, c) in filled_cells:
                parts.append(f"[{coord}]='{filled_cells[(r, c)]}'")
            else:
                parts.append(f"[{coord}]=EMPTY")
        lines.append(" ".join(parts))
        
    return "\n".join(lines), protected_cells


# ──── Deterministic vertical-block repeating-table detector (no AI call needed) ────

FIELD_KEYWORDS = {
    "client_name": ["client name", "client referance", "client reference", "customer name", "end user"],
    "project_name": ["project name", "name of project"],
    "location": ["location"],
    "area_sqft": ["area (sqft)", "area sqft", "area"],
    "amount": ["project cost", "contract value", "project value", "value in lacs", "value"],
    "duration": ["duration"],
    "scope_of_work": ["scope of work", "scope"],
    "contact_name": ["contact person name", "concern person", "contact person"],
    "contact_designation": ["designation"],
    "contact_phone": ["contact details", "contact no", "phone", "mobile"],
    "contact_email": ["email"],
    "start_date": ["start date"],
    "completion_date": ["completion date", "end date"],
}


def _guess_field_key(sub_label: str):
    low = sub_label.lower()
    for key, kws in FIELD_KEYWORDS.items():
        for kw in kws:
            if kw in low:
                return key
    return None


def find_field_value_in_record(rec_data: dict, field_key: str, column_labels: dict | None = None) -> str | None:
    """
    Given one imported project-data row (rec_data: {normalized_column_key: value})
    find the value for a canonical field like "contact_name" or "project_name",
    even when the source spreadsheet's own column header uses completely
    different wording (real imports are inconsistent -- "Client" vs "Client
    Name" vs "NAME OF CLIENT" vs "Party Name", or a single merged column like
    "Contact reference (Name / Organisation / Email ID / Phone Number)").

    FIX (P0 -- project-detail table cells silently staying blank): the
    previous version matched a short hardcoded keyword list as a plain
    substring against the normalized snake_case column KEY (e.g. "client" in
    "client_name"), and for several fields -- client_name, project_name,
    contact_name, contact_designation, contact_phone, contact_email -- there
    was no keyword list at all, so it fell back to literally requiring the
    key to contain the field name itself (e.g. "client_name" in "client_name")
    -- which only matches a source column already named almost exactly that.
    Any other real-world header ("Client", "Customer", "NAME OF CLIENT")
    silently produced no value, even though a project WAS selected and DID
    have the data. This scores every column's REAL display label (via
    `column_labels`, sourced from ProjectDataColumn -- falls back to a
    prettified version of the key if not available) against the field's known
    phrasings (FIELD_KEYWORDS) with the same fuzzy matcher used everywhere
    else in this codebase, instead of requiring an exact substring hit.
    """
    try:
        from rapidfuzz import fuzz
    except ImportError:
        fuzz = None

    keywords = FIELD_KEYWORDS.get(field_key, [field_key.replace("_", " ")])

    def norm(s):
        return re.sub(r'[^a-z0-9]+', ' ', str(s or '').lower()).strip()

    normalized_keywords = [norm(kw) for kw in keywords]

    best_value, best_score, second_best_score = None, 0, 0
    for col_key, value in (rec_data or {}).items():
        if not value or not isinstance(value, str) or not value.strip():
            continue
        label = (column_labels or {}).get(col_key) or col_key.replace('_', ' ')
        candidate = norm(label)
        if not candidate:
            continue

        if fuzz:
            score = max(fuzz.token_set_ratio(kw, candidate) for kw in normalized_keywords)
        else:
            score = 100 if any(kw in candidate or candidate in kw for kw in normalized_keywords) else 0

        # Length-aware guard (same pattern as _match_company_context_value):
        # short candidate labels (<=2 tokens, e.g. "Project Name") are exactly
        # where token_set_ratio breaks down -- a keyword phrase sharing just
        # ONE generic word ("project value" vs "project name" scores 80) can
        # still clear a flat threshold, so short candidates require a
        # near-exact match instead of a merely-similar one.
        candidate_tokens = candidate.split()
        required = 92 if len(candidate_tokens) <= 2 else 82

        if score >= required:
            value_clean = value.strip()
            if score > best_score:
                second_best_score = best_score
                best_score = score
                best_value = value_clean
            elif score > second_best_score and value_clean != best_value:
                second_best_score = score

    # Ambiguity guard: if a second, different column scored nearly as well as
    # the winner, this is a coin flip -- refuse to guess rather than risk
    # filling the wrong data into the form.
    if best_value is not None and (best_score - second_best_score) < 8:
        return None

    if best_value:
        return best_value

    # FIX: many real project trackers record Start Date / End Date but never
    # a separate "Duration" figure -- rather than leave it blank when those
    # two dates ARE present and parseable, compute it.
    if field_key == "duration":
        return _compute_duration_from_dates(rec_data, column_labels)

    return None


def _compute_duration_from_dates(rec_data: dict, column_labels: dict | None = None) -> str | None:
    """Fallback for a missing "duration" column: derive it from whichever
    columns best match start_date/completion_date instead. Returns e.g.
    "13 months" or "18 days", or None if either date is missing/unparsable."""
    start_val = find_field_value_in_record(rec_data, "start_date", column_labels)
    end_val = find_field_value_in_record(rec_data, "completion_date", column_labels)
    if not start_val or not end_val:
        return None

    def parse(s):
        s = str(s).strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        return None

    start_dt, end_dt = parse(start_val), parse(end_val)
    if not start_dt or not end_dt or end_dt < start_dt:
        return None

    days = (end_dt - start_dt).days
    months = round(days / 30.44)
    if months >= 1:
        return f"{months} month{'s' if months != 1 else ''}"
    return f"{days} day{'s' if days != 1 else ''}"


def _col_letter(idx: int) -> str:
    name = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        name = chr(65 + rem) + name
    return name


def detect_vertical_repeating_blocks(ws) -> list[dict]:
    label_merges = []
    for mr in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = mr.bounds
        if max_row > min_row:
            anchor = ws.cell(row=min_row, column=min_col)
            if anchor.value:
                label_merges.append({
                    "min_row": min_row, "max_row": max_row,
                    "min_col": min_col, "max_col": max_col,
                    "text": str(anchor.value).strip()
                })

    def strip_trailing_number(s: str) -> str:
        return re.sub(r'[\s\-\u2013_:]*\d+\s*$', '', s).strip().lower()

    groups: dict = {}
    for lm in label_merges:
        key = strip_trailing_number(lm["text"])
        if key:
            groups.setdefault(key, []).append(lm)

    detected = []
    for key, members in groups.items():
        if len(members) < 2: continue
        members.sort(key=lambda m: m["min_row"])
        rows_per_block = members[0]["max_row"] - members[0]["min_row"] + 1
        sub_label_col = members[0]["max_col"] + 1
        sub_labels = {}
        for offset in range(rows_per_block):
            r = members[0]["min_row"] + offset
            cell = ws.cell(row=r, column=sub_label_col)
            if cell.value:
                sub_labels[offset] = str(cell.value).strip()

        field_row_offsets = {}
        for offset, sub_label in sub_labels.items():
            fk = _guess_field_key(sub_label)
            if fk and fk not in field_row_offsets:
                field_row_offsets[fk] = offset
        
        answer_col_idx = sub_label_col + 1
        contact_ish = {"contact_name", "contact_phone", "contact_email", "contact_designation"}
        project_ish = {"amount", "scope_of_work", "project_name", "duration", "area_sqft", "location"}
        
        if any(k in field_row_offsets for k in project_ish):
            table_type = "project_details"
        else:
            table_type = "project_reference"

        detected.append({
            "subheading": members[0]["text"],
            "table_type": table_type,
            "layout": "vertical",
            "block_start_rows": [m["min_row"] for m in members],
            "rows_per_block": rows_per_block,
            "answer_column": _col_letter(answer_col_idx),
            "field_row_offsets": field_row_offsets,
        })
    return detected


def excel_to_all_sheet_maps(file_bytes: bytes) -> tuple[dict, dict, dict, dict]:
    wb_values = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    wb_formulas = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    maps = {}
    protected = {}
    states = {}
    vertical_blocks = {}
    for sheet_name in wb_values.sheetnames:
        ws = wb_values[sheet_name]
        ws_formulas = wb_formulas[sheet_name] if sheet_name in wb_formulas.sheetnames else None
        m, p = build_sheet_cell_map(ws, ws_formulas)
        maps[sheet_name] = m
        protected[sheet_name] = p
        states[sheet_name] = ws.sheet_state
        try:
            vertical_blocks[sheet_name] = detect_vertical_repeating_blocks(ws)
        except Exception as e:
            print(f"[WARN] Vertical-block detection failed for '{sheet_name}': {e}")
            vertical_blocks[sheet_name] = []
    return maps, protected, states, vertical_blocks


# Keep old name working for anything else that imports it (single-sheet, first sheet)
def excel_to_cell_map(file_bytes: bytes) -> tuple[str, str]:
    wb_values = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    wb_formulas = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    ws = wb_values.active
    ws_formulas = wb_formulas.active
    m, _ = build_sheet_cell_map(ws, ws_formulas)
    return ws.title, m



SECTION_KEYWORDS = {
    "General Information": (
        "general information", "company information", "basic information", "company profile",
        "organization details", "organisation details", "contact details", "statutory details",
        "vendor details", "supplier details", "contractor details"
    ),
    "Financial Info": (
        "financial", "finance", "turnover", "annual turnover", "balance sheet", "p&l",
        "profit", "loss", "net worth", "assets", "liabilities", "working capital",
        "solvency", "banker", "income tax", "itr", "revenue"
    ),
    "Customer References": (
        "customer reference", "customer references", "client reference", "client references",
        "reference details", "client contact", "contact person"
    ),
    "Major Work Done / Projects Done": (
        "major work done", "major works done", "major projects", "projects done",
        "project details", "project history", "past projects", "completed projects",
        "ongoing projects", "past performance", "experience record", "work done",
        "scope of work", "project cost", "contract value"
    ),
    "Certificates / Licenses / Attachments": (
        "certificate", "certificates", "license", "licenses", "licence", "licences",
        "attachment", "attachments", "documents", "document list", "enclosure",
        "annexure", "upload", "attached"
    ),
}


def infer_company_section(text: str, default: str = "General Information") -> str:
    low = (text or "").lower()
    for section, keywords in SECTION_KEYWORDS.items():
        if any(keyword in low for keyword in keywords):
            return section
    return default


def build_workbook_form_json(file_bytes: bytes, sheet_maps: dict | None = None) -> dict:
    """
    Build a real JSON structure from the uploaded Excel workbook. This complements
    the compact text cell map: it keeps the sheet/section/field nesting so both
    deterministic matching and GPT can search the right company-data section first.
    """
    if sheet_maps is None:
        sheet_maps, _, _, _ = excel_to_all_sheet_maps(file_bytes)

    workbook = {
        "format": "preque_openpyxl_form_structure_v1",
        "lookup_order": [
            "Use the field's section first",
            "If missing, search every other company-data section before leaving blank",
        ],
        "sheets": {},
    }

    for sheet_name, cell_map in sheet_maps.items():
        sections = {}
        current_section = "General Information"
        row_tokens = {}
        for coord, val in re.findall(r"\[([A-Z]+\d+)\]=(EMPTY|'[^']*')", cell_map):
            row = int(re.search(r"\d+", coord).group(0))
            if val.startswith("'") and val.endswith("'"):
                val = val[1:-1]
            row_tokens.setdefault(row, []).append({"cell": coord, "value": val})

        candidates = {}
        try:
            from services.field_matcher import extract_candidates_from_map
            candidates = extract_candidates_from_map(cell_map)
        except Exception:
            candidates = {}

        for row in sorted(row_tokens):
            tokens = row_tokens[row]
            filled_texts = [t["value"] for t in tokens if t["value"] and t["value"] != "EMPTY"]
            row_text = " ".join(filled_texts)
            detected_section = infer_company_section(row_text, current_section)
            if detected_section != current_section and filled_texts:
                current_section = detected_section

            section_obj = sections.setdefault(current_section, {
                "name": current_section,
                "rows": [],
                "fillable_fields": [],
            })
            section_obj["rows"].append({
                "row": row,
                "text": row_text,
                "cells": tokens[:],
            })

            for token in tokens:
                if token["value"] != "EMPTY":
                    continue
                labels = candidates.get(token["cell"], [])
                local_text = " ".join(labels + filled_texts)
                field_section = infer_company_section(local_text, current_section)
                sections.setdefault(field_section, {
                    "name": field_section,
                    "rows": [],
                    "fillable_fields": [],
                })["fillable_fields"].append({
                    "cell": token["cell"],
                    "labels": labels,
                    "row_context": row_text,
                    "section": field_section,
                })

        workbook["sheets"][sheet_name] = {
            "sheet_name": sheet_name,
            "sections": list(sections.values()),
        }

    return workbook


def _form_fields_by_cell(sheet_form_json: dict | None) -> dict:
    fields = {}
    if not sheet_form_json:
        return fields
    for section in sheet_form_json.get("sections", []):
        for field in section.get("fillable_fields", []):
            cell = field.get("cell")
            if cell:
                fields[cell] = field
    return fields


def _is_attachment_only_request(labels: list[str], section: str = "") -> bool:
    text = " ".join([section] + [str(label) for label in labels if label]).lower()
    if not text:
        return False
    asks_for_attachment = any(token in text for token in [
        "attach", "attachment", "attached", "enclosed", "enclosure", "upload",
        "submit", "provide copy", "copy of", "supporting document"
    ])
    mentions_document = any(token in text for token in [
        "certificate", "license", "licence", "registration", "document", "annexure"
    ])
    asks_for_identifier = any(token in text for token in [
        "number", "no.", "no ", "date", "valid", "expiry", "issued", "authority",
        "name of", "details of", "gst no", "pan no",
        # FIX (P0): "No. of X" / "Number of X" / "How many" rows ask for a COUNT,
        # not a document -- e.g. "No. of projects where bonus is received" was
        # getting misfired into "Attached" because it mentions neither "number"
        # verbatim in a way the old check caught, nor did it get excluded properly.
        "no of", "number of", "how many", "count of",
    ])
    return (asks_for_attachment or mentions_document) and not asks_for_identifier


# FIX (P0 -- wrong financial metric / wrong year): deterministic keyword groups.
# The previous version scored every FinancialRecord with fuzz.token_set_ratio plus
# a chain of "if/elif" keyword special-cases, so if the *correct* record (e.g. the
# one whose metric_label actually contains "profit") had a lower base fuzzy score
# than expected, or its value happened to be stored differently than assumed, an
# unrelated metric (e.g. "Fixed Assets") could still out-score it once the +35
# year-match bonus was added on top of a merely-average base score. That's exactly
# how "Profit of firm" ended up filled with Fixed Assets figures. This replaces all
# of that with plain substring containment against explicit keyword groups -- no
# fuzzy score, no bonus stacking, no possibility of a Fixed-Assets record ever
# satisfying a "profit" query. The fiscal year requirement is now mandatory (not a
# scoring bonus): if the label mentions a year and no record for that exact metric
# and year exists, the cell is left blank rather than filled from the wrong year.
FINANCIAL_METRIC_KEYWORD_GROUPS = [
    ["turnover", "turn over", "revenue"],
    ["net profit", "profit after tax", "profit"],
    ["net worth", "networth", "shareholders fund", "shareholder fund"],
    ["total assets", "fixed assets", "current assets", "assets"],
    ["total liabilities", "current liabilities", "liabilities"],
    ["working capital"],
]


def _match_financial_record(labels: list[str], section: str, db: Session) -> str | None:
    from models.database import FinancialRecord

    text = " ".join([section] + [str(label) for label in labels if label]).lower()
    if not text:
        return None

    def norm(s: str) -> str:
        return re.sub(r'[^a-z0-9]+', ' ', str(s or '').lower()).strip()

    norm_text = norm(text)

    financial_hint = infer_company_section(text) == "Financial Info"
    financial_hint = financial_hint or any(
        kw in norm_text for group in FINANCIAL_METRIC_KEYWORD_GROUPS for kw in group
    )
    if not financial_hint:
        return None

    # Which keyword group is this label actually asking about? First group that has
    # any keyword present in the label wins -- groups are ordered most-specific-first
    # ("net profit" before generic "assets") to avoid a generic word claiming a more
    # specific field.
    target_group = None
    for group in FINANCIAL_METRIC_KEYWORD_GROUPS:
        if any(kw in norm_text for kw in group):
            target_group = group
            break
    if not target_group:
        return None

    def record_matches_metric(record) -> bool:
        metric_text = norm(record.metric_label or record.metric_key or "")
        return any(kw in metric_text for kw in target_group)

    all_records = db.query(FinancialRecord).all()
    candidates = [r for r in all_records if r.value and record_matches_metric(r)]
    if not candidates:
        return None

    has_year_hint = bool(re.search(r'\b(?:fy\s*)?\d{2,4}\s*[-/]\s*\d{2,4}\b', text, re.IGNORECASE))

    if has_year_hint:
        # A year was mentioned -- ONLY accept a record for that exact fiscal year.
        # Never fall back to a different year's figure for a year-specific cell.
        for record in candidates:
            for variant in _fiscal_year_variants(record.fiscal_year or ""):
                if norm(variant) and norm(variant) in norm_text:
                    unit_str = f" {record.unit}" if record.unit else ""
                    return f"{record.value}{unit_str}"
        return None

    # No year mentioned -- only safe to answer if there's exactly one candidate for
    # this metric across the whole company; otherwise it's genuinely ambiguous.
    if len(candidates) == 1:
        record = candidates[0]
        unit_str = f" {record.unit}" if record.unit else ""
        return f"{record.value}{unit_str}"
    return None


def _match_company_context_value(labels: list[str], section: str, company_context: dict) -> str | None:
    """
    FIX (P0 -- "DIN everywhere" bug): the previous version fuzzy-matched
    `section_name + " " + label` against `section_name + " " + leaf_key` for every
    candidate. Because every field in the same section shares those section-name
    tokens on both sides, token_set_ratio scores were dominated by the shared
    boilerplate rather than the actual field text -- a short, generic leaf like
    "DIN" ends up looking MORE similar to almost any query in its section than the
    actually-correct field does (empirically: DIN scored 90.5 vs. Company Name's
    85.7 for the query "name of the firm"). The fix: never blend the section name
    into the scored strings. Section is used only to decide which section's fields
    to search FIRST (an ordering hint), never as text that inflates the score. A
    length-aware threshold also guards against short candidate strings (like "DIN",
    "PAN", "GST") winning on fuzzy noise alone -- short candidates now require a
    near-exact match, not just a generically "similar-length" one.
    """
    try:
        from rapidfuzz import fuzz
    except Exception:
        fuzz = None

    def norm(text: str) -> str:
        return re.sub(r'[^a-z0-9]+', ' ', str(text or '').lower()).strip()

    # Query is built ONLY from the actual label text -- never include the section name here.
    query = norm(" ".join([str(label) for label in labels if label]))
    if not query:
        return None

    def iter_leaves(value, key_path: str = ""):
        """Yields (leaf_key_text, leaf_value) pairs. key_path tracks only the
        immediate leaf key (e.g. 'DIN'), never accumulates the section name, so the
        candidate string fuzzy-matched against `query` is just the field's own label."""
        if isinstance(value, dict):
            for key, nested in value.items():
                yield from iter_leaves(nested, key)
        elif isinstance(value, list):
            for item in value:
                yield from iter_leaves(item, key_path)
        else:
            leaf_value = str(value or "").strip()
            if leaf_value:
                yield key_path, leaf_value

    best_value = None
    best_score = 0
    second_best_score = 0
    section_order = []
    if section and section in company_context:
        section_order.append(section)
    section_order.extend([name for name in company_context.keys() if name not in section_order and name != "Lookup Guidance"])

    for section_name in section_order:
        section_data = company_context.get(section_name)
        for leaf_key, leaf_value in iter_leaves(section_data, ""):
            candidate = norm(leaf_key)
            if not candidate:
                continue
            score = fuzz.token_set_ratio(query, candidate) if fuzz else (100 if candidate in query or query in candidate else 0)

            # Length-aware guard: short candidate strings (<=2 tokens after
            # normalization, e.g. "din", "pan no", "gst") are exactly where fuzzy
            # scoring breaks down -- require them to be near-exact, not just
            # superficially similar in length/shared words.
            candidate_tokens = candidate.split()
            required = 92 if len(candidate_tokens) <= 2 else 84

            if score >= required:
                if score > best_score:
                    second_best_score = best_score
                    best_score = score
                    best_value = leaf_value
                elif score > second_best_score and leaf_value != best_value:
                    second_best_score = score

        # If we already found a strong match in the section the field actually
        # belongs to, don't let a later section's coincidental match override it.
        if section_name == section and best_value is not None:
            break

    # FIX (P1 -- ambiguous generic queries like bare "Name"): if a second, DIFFERENT
    # candidate scored nearly as well as the winner, this query is genuinely
    # ambiguous (e.g. "Name" alone could mean "Company Name" or "Contact Person
    # Name") -- refuse to guess. Leaving it blank for GPT-4o vision (which has
    # visual/positional context this deterministic layer doesn't) is safer than
    # confidently answering with a coin-flip.
    if best_value is not None and (best_score - second_best_score) < 8:
        return None

    return best_value

def _find_libreoffice() -> str:
    import shutil as _shutil, platform, glob
    lo = _shutil.which("libreoffice") or _shutil.which("soffice")
    if lo: return lo
    if platform.system() == "Windows":
        candidates = glob.glob(r"C:\Program Files\LibreOffice*\program\soffice.exe") + \
                     glob.glob(r"C:\Program Files (x86)\LibreOffice*\program\soffice.exe")
        for p in candidates:
            if os.path.exists(p): return p
    elif platform.system() == "Darwin":
        for p in ["/Applications/LibreOffice.app/Contents/MacOS/soffice",
                  "/opt/homebrew/bin/libreoffice"]:
            if os.path.exists(p): return p
    raise RuntimeError(
        "LibreOffice not found. Install from https://libreoffice.org\n"
        "Or set LIBREOFFICE_PATH in your .env file."
    )


def excel_to_images_per_sheet(file_bytes: bytes, filename: str) -> tuple[dict, str]:
    """
    Render each sheet of the Excel workbook to its own PNG image(s) via LibreOffice + pdf2image.
    Returns ({sheet_name: [image_paths]}, tmpdir)

    Strategy: split the workbook into one temp .xlsx per sheet (openpyxl copy),
    convert each to PDF individually so page-to-sheet mapping is exact, then rasterize.
    """
    import platform
    lo_path = os.getenv("LIBREOFFICE_PATH") or _find_libreoffice()
    tmpdir = tempfile.mkdtemp()

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    sheet_names = wb.sheetnames

    lo_profile = os.path.join(tmpdir, "lo_profile")
    os.makedirs(lo_profile, exist_ok=True)
    profile_uri = "file:///" + lo_profile.replace("\\", "/")

    env = os.environ.copy()
    if platform.system() == "Windows":
        env["PATH"] = os.path.dirname(lo_path) + os.pathsep + env.get("PATH", "")

    images_by_sheet = {}

    for sheet_name in sheet_names:
        # Build a single-sheet workbook copy so LibreOffice renders exactly this tab.
        single_wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
        for other in list(single_wb.sheetnames):
            if other != sheet_name:
                del single_wb[other]

        # Ensure the sheet is visible before saving, as a workbook cannot have only hidden sheets
        single_wb[sheet_name].sheet_state = 'visible'

        safe_sheet = re.sub(r'[^\w\-]', '_', sheet_name)
        xlsx_path = os.path.join(tmpdir, f"{safe_sheet}.xlsx")
        single_wb.save(xlsx_path)
        single_wb.close()

        sheet_outdir = os.path.join(tmpdir, safe_sheet)
        os.makedirs(sheet_outdir, exist_ok=True)

        result = subprocess.run(
            [lo_path, "--headless", "--norestore", "--nofirststartwizard",
             f"-env:UserInstallation={profile_uri}",
             "--convert-to", "pdf", "--outdir", sheet_outdir, xlsx_path],
            capture_output=True, text=True, timeout=120, env=env
        )
        if result.returncode != 0:
            print(f"[WARN] LibreOffice failed for sheet '{sheet_name}': {result.stderr or result.stdout}")
            images_by_sheet[sheet_name] = []
            continue

        pdfs = [f for f in os.listdir(sheet_outdir) if f.endswith(".pdf")]
        if not pdfs:
            images_by_sheet[sheet_name] = []
            continue

        from pdf2image import convert_from_path
        pages = convert_from_path(os.path.join(sheet_outdir, pdfs[0]), dpi=150)
        paths = []
        for i, page in enumerate(pages):
            p = os.path.join(sheet_outdir, f"page_{i}.png")
            page.save(p, "PNG")
            paths.append(p)
        images_by_sheet[sheet_name] = paths

    wb.close()
    return images_by_sheet, tmpdir


def _default_sheet_analysis() -> dict:
    return {
        "classification": "FILLABLE",
        "fillable_columns": [],
        "reserved_columns": [],
        "table_info": {"is_project_table": False},
    }


def analyze_sheet(sheet_name: str, cell_map: str) -> dict:
    """
    Single GPT-4o-mini call answering all three per-sheet questions that used to
    be three separate round trips (classification, column policy, project-table
    detection) -- see the FIX note above SHEET_ANALYSIS_PROMPT. Returns:
    {"classification": "FILLABLE"|"INFO_ONLY", "fillable_columns": [...],
     "reserved_columns": [...], "table_info": {...}}
    On any failure, falls back to the same safe defaults the three separate
    functions used to return individually (FILLABLE, no column restriction,
    not a project table) so a flaky call degrades gracefully instead of
    silently dropping a sheet.
    """
    if not cell_map.strip():
        return {**_default_sheet_analysis(), "classification": "INFO_ONLY"}
    try:
        response = openai_client.chat.completions.create(
            model=MINI_MODEL,
            # Gemini 3.x models can't have internal "thinking" disabled and
            # spend part of the token budget on it before emitting the
            # visible JSON -- reasoning_effort="low" keeps that spend down for
            # what's a mechanical classification task, and max_tokens is
            # generous so a bigger-than-expected reasoning pass still leaves
            # room for the actual JSON instead of truncating it mid-string.
            max_tokens=8000,
            reasoning_effort="low",
            response_format={"type": "json_object"},
            messages=[{
                "role": "user",
                "content": SHEET_ANALYSIS_PROMPT.format(
                    sheet_name=sheet_name,
                    cell_map=cell_map[:6000]  # keep the call cheap
                )
            }]
        )
        raw = response.choices[0].message.content.strip()
        data = json.loads(raw)
        classification = str(data.get("classification", "FILLABLE")).strip().upper()
        columns = data.get("columns") or {}
        return {
            "classification": "FILLABLE" if classification != "INFO_ONLY" else "INFO_ONLY",
            "fillable_columns": columns.get("fillable_columns", []),
            "reserved_columns": columns.get("reserved_columns", []),
            "table_info": data.get("project_table") or {"is_project_table": False},
        }
    except Exception as e:
        print(f"[WARN] Sheet analysis failed for '{sheet_name}': {e} — defaulting to FILLABLE / no column restriction / no project table")
        return _default_sheet_analysis()



PROJECT_TABLE_FIELDS = {
    "project_name", "client_name", "location", "area_sqft", "amount",
    "start_date", "completion_date", "duration", "scope_of_work",
    "contact_name", "contact_designation", "contact_phone", "contact_email"
}


def normalize_detected_table_info(table_info: dict, cell_map: str) -> dict:
    if not table_info or not table_info.get("is_project_table"):
        return table_info or {"is_project_table": False}

    mapping = table_info.get("mapping") or {}
    normalized = {}
    for key, value in mapping.items():
        key_s = str(key or "").strip()
        val_s = str(value or "").strip()
        if re.match(r'^[A-Z]{1,3}$', key_s) and val_s:
            guessed = _guess_field_key(val_s)
            if guessed:
                normalized[guessed] = key_s
        elif key_s in PROJECT_TABLE_FIELDS and re.match(r'^[A-Z]{1,3}$', val_s):
            normalized[key_s] = val_s

    if normalized:
        table_info["mapping"] = normalized

    subheading_text = str(table_info.get("subheading") or "").lower()
    mapped_keys = set((table_info.get("mapping") or {}).keys())
    project_markers = SECTION_KEYWORDS["Major Work Done / Projects Done"]
    reference_markers = SECTION_KEYWORDS["Customer References"]
    project_keys = {"project_name", "location", "area_sqft", "amount", "duration", "scope_of_work", "start_date", "completion_date"}
    contact_keys = {"contact_name", "contact_designation", "contact_phone", "contact_email"}

    if any(marker in subheading_text for marker in project_markers):
        table_info["table_type"] = "project_details"
    elif any(marker in subheading_text for marker in reference_markers):
        table_info["table_type"] = "project_reference"
    elif len(mapped_keys & project_keys) >= len(mapped_keys & contact_keys):
        table_info["table_type"] = "project_details"

    return table_info

def ai_fill_sheet(
    sheet_name: str,
    b64_images: list[str],
    cell_map: str,
    company_context: dict,
    protected_cells: set[str] = None,
    reserved_cols: list[str] = None,
    fillable_cols: list[str] = None,
    sheet_form_json: dict | None = None
) -> dict:
    """Fill ONE sheet using its own image(s) + its own cell map. Returns {cell_address: value}."""
    if not b64_images or not cell_map.strip():
        return {}

    all_fills = {}
    seen_cells = set()

    for i, b64 in enumerate(b64_images):
        content = [
            {
                "type": "image_url",
                "image_url": {"url": f"data:image/png;base64,{b64}", "detail": "high"}
            },
            {
                "type": "text",
                "text": (
                    f"SHEET NAME: {sheet_name}\n\n"
                    f"CELL MAP (exact coordinates for every cell on THIS SHEET):\n"
                    f"{cell_map}\n\n"
                    f"OPENPYXL FORM STRUCTURE JSON (sections and fillable cells for THIS SHEET):\n"
                    f"{json.dumps(sheet_form_json or {}, indent=2)}\n\n"
                    f"COMPANY DATA JSON (already grouped by section):\n{json.dumps(company_context, indent=2)}\n\n"
                    "Look at the image to understand this sheet visually. "
                    "Use the cell map to get exact cell coordinates for THIS SHEET ONLY. "
                    "Return JSON mapping cell addresses to values for ALL fillable fields on this sheet."
                )
            }
        ]
        try:
            response = openai_client.chat.completions.create(
                model=VISION_MODEL,
                # A real sheet's prompt (image + full cell map + full company
                # context JSON) triggers much heavier internal reasoning than
                # a simple test call did -- 6000 wasn't enough and truncated
                # mid-JSON (json.loads failing with "Unterminated string").
                # reasoning_effort="low" keeps that reasoning pass from
                # ballooning further, and max_tokens has real headroom so an
                # underestimate doesn't silently cut the response off again.
                max_tokens=24000,
                reasoning_effort="low",
                response_format={"type": "json_object"},
                messages=[
                    {"role": "system", "content": FILL_SYSTEM_PROMPT},
                    {"role": "user", "content": content}
                ]
            )
            raw = response.choices[0].message.content.strip()
            if raw.startswith("```"):
                raw = raw.split("\n", 1)[1].rsplit("```", 1)[0].strip()

            page_fills = json.loads(raw)
            new_fills = {}
            for k, v in page_fills.items():
                k = k.strip()
                if not re.match(r'^[A-Z]{1,3}\d+$', k) or not v or k in seen_cells:
                    continue
                if protected_cells and k in protected_cells:
                    continue
                
                col_letter = re.match(r'^([A-Z]+)', k).group(1)
                if reserved_cols and col_letter in reserved_cols:
                    continue
                if fillable_cols and reserved_cols and col_letter not in fillable_cols:
                    continue
                    
                new_fills[k] = str(v)
                
            seen_cells.update(new_fills.keys())
            all_fills.update(new_fills)
            print(f"  [GPT-4o] Sheet '{sheet_name}' page {i+1}: filled {len(new_fills)} cells")

        except json.JSONDecodeError as e:
            print(f"  [GPT-4o] Sheet '{sheet_name}' page {i+1}: JSON parse error — {e}")
        except Exception as e:
            print(f"  [GPT-4o] Sheet '{sheet_name}' page {i+1}: Error — {e}")

    return all_fills


def ai_fill_workbook(
    file_bytes: bytes,
    filename: str,
    company_context: dict,
    db: Session,
    workbook_form_json: dict | None = None
) -> tuple[dict, dict, str, dict, dict, list]:
    from services.field_matcher import FieldMatcher, extract_candidates_from_map
    try:
        matcher = FieldMatcher(db)
    except Exception as e:
        print("Could not load matcher:", e)
        matcher = None

    sheet_maps, protected_cells_by_sheet, sheet_states, vertical_blocks_by_sheet = excel_to_all_sheet_maps(file_bytes)
    if workbook_form_json is None:
        workbook_form_json = build_workbook_form_json(file_bytes, sheet_maps)

    sheet_status = {}
    sheet_column_policies = {}
    sheet_table_info = {}

    sheets_to_analyze = []
    for sheet_name, cmap in sheet_maps.items():
        state = sheet_states.get(sheet_name, 'visible')
        if state in ('hidden', 'veryHidden'):
            sheet_status[sheet_name] = 'INFO_ONLY'
            print(f"[Classify] '{sheet_name}' is {state} -> skipping classification, marking INFO_ONLY")
        else:
            sheets_to_analyze.append((sheet_name, cmap))

    # FIX (P1 -- request-timeout / slow-UX risk): these per-sheet analysis calls
    # are independent of each other and touch no shared DB/session state, so run
    # them concurrently instead of one-by-one -- for a multi-tab form this turns
    # N sequential round trips into roughly one round trip's worth of wall time.
    if sheets_to_analyze:
        with ThreadPoolExecutor(max_workers=min(8, len(sheets_to_analyze))) as executor:
            future_to_sheet = {
                executor.submit(analyze_sheet, sheet_name, cmap): sheet_name
                for sheet_name, cmap in sheets_to_analyze
            }
            for future in as_completed(future_to_sheet):
                sheet_name = future_to_sheet[future]
                analysis = future.result()
                status = analysis["classification"]
                sheet_status[sheet_name] = status
                print(f"[Classify] '{sheet_name}' -> {status}")
                if status == "FILLABLE":
                    sheet_column_policies[sheet_name] = {
                        "fillable": analysis["fillable_columns"],
                        "reserved": analysis["reserved_columns"],
                    }
                    sheet_table_info[sheet_name] = analysis["table_info"]
                    print(f"[Classify Columns] '{sheet_name}' -> fillable: {analysis['fillable_columns']}, reserved: {analysis['reserved_columns']}")

    # Preserve original workbook sheet order (thread completion order is
    # nondeterministic) for all downstream processing.
    fillable_sheets = [sn for sn in sheet_maps if sheet_status.get(sn) == "FILLABLE"]

    if not fillable_sheets:
        return {}, sheet_status, "", {}, {}, []

    images_by_sheet, tmpdir = excel_to_images_per_sheet(file_bytes, filename)

    all_fills = {}
    fill_sources = {}
    match_summary = {"alias_matched": 0, "ai_guessed": 0, "sheets_skipped_vision_call": []}
    pending_project_tables = []

    try:
        for sheet_name in fillable_sheets:
            cmap = sheet_maps[sheet_name]
            resolved_cells = {}
            policy = sheet_column_policies.get(sheet_name, {})
            reserved_cols = policy.get("reserved", [])
            fillable_cols = policy.get("fillable", [])
            sheet_form_json = (workbook_form_json or {}).get("sheets", {}).get(sheet_name, {})
            fields_by_cell = _form_fields_by_cell(sheet_form_json)
            
            def is_allowed_col(coord):
                col_letter = re.match(r'^([A-Z]+)', coord).group(1)
                if reserved_cols and col_letter in reserved_cols:
                    return False
                if fillable_cols and reserved_cols and col_letter not in fillable_cols:
                    return False
                return True

            if sheet_name not in protected_cells_by_sheet:
                protected_cells_by_sheet[sheet_name] = set()

            vblocks = vertical_blocks_by_sheet.get(sheet_name, [])
            for vb in vblocks:
                print(f"  [Vertical Table Mode] '{sheet_name}' section '{vb['subheading']}' -> "
                      f"{len(vb['block_start_rows'])} repeating blocks detected deterministically "
                      f"(type={vb['table_type']})")
                pending_project_tables.append({
                    "sheet_name": sheet_name,
                    "table_type": vb["table_type"],
                    "subheading": vb["subheading"],
                    "layout": "vertical",
                    "answer_column": vb["answer_column"],
                    "block_start_rows": vb["block_start_rows"],
                    "rows_per_block": vb["rows_per_block"],
                    "field_row_offsets": vb["field_row_offsets"],
                    "available_row_count": len(vb["block_start_rows"]),
                    "max_rows": None,
                })
                for block_start in vb["block_start_rows"]:
                    for offset in range(vb["rows_per_block"]):
                        protected_cells_by_sheet[sheet_name].add(f"{vb['answer_column']}{block_start + offset}")

            if not vblocks:
                # Reuse the project-table detection already produced by the
                # combined analyze_sheet() call above -- no second AI call needed.
                table_info = normalize_detected_table_info(sheet_table_info.get(sheet_name, {"is_project_table": False}), cmap)
                if table_info.get("is_project_table"):
                    print(f"  [Table Mode] '{sheet_name}' is a Project Table. Adding to pending_project_tables.")
                    start_row = table_info.get("start_row", 1)
                    mapping = table_info.get("mapping", {})
                    
                    available_rows = []
                    first_col = next(iter(mapping.values())) if mapping else None
                    if first_col:
                        r = start_row
                        while f"[{first_col}{r}]=EMPTY" in cmap:
                            available_rows.append(r)
                            r += 1
                    
                    pending_project_tables.append({
                        "sheet_name": sheet_name,
                        "table_type": table_info.get("table_type", "project_details"),
                        "subheading": table_info.get("subheading"),
                        "start_row": start_row,
                        "mapping": mapping,
                        "available_row_count": len(available_rows),
                        "max_rows": table_info.get("max_rows")
                    })
                    
                    for r in available_rows:
                        for c_letter in mapping.values():
                            protected_cells_by_sheet[sheet_name].add(f"{c_letter}{r}")

            if matcher:
                candidates = extract_candidates_from_map(cmap)
                for coord, labels in candidates.items():
                    if coord in protected_cells_by_sheet.get(sheet_name, set()):
                        continue
                    if not is_allowed_col(coord):
                        continue
                    field_context = fields_by_cell.get(coord, {})

                    # FIX (P0 -- "DIN everywhere" bug, path 2): the section name used
                    # to be smuggled into context_labels itself (not just passed as
                    # the separate `section` arg), which re-contaminates the fuzzy
                    # query even after _match_company_context_value stopped
                    # concatenating `section` internally. Also strip any residual
                    # "--- Section: ... ---" banner text that can leak in via
                    # row_context when a fillable cell shares a row with a header.
                    def _clean_context_text(v):
                        s = str(v or "").strip()
                        if not s or (s.startswith("--- Section:") and s.endswith("---")):
                            return ""
                        return s

                    raw_context_labels = (
                        labels +
                        field_context.get("labels", []) +
                        [field_context.get("row_context", "")]
                        # NOTE: field_context["section"] is deliberately NOT included
                        # here -- it's passed separately as `section` below, used only
                        # for search ordering, never blended into the matched text.
                    )
                    context_labels = list(dict.fromkeys(
                        cleaned for cleaned in (_clean_context_text(l) for l in raw_context_labels) if cleaned
                    ))
                    section = field_context.get("section", "")

                    if _is_attachment_only_request(context_labels, section):
                        resolved_cells[coord] = "Attached"
                        fill_sources[f"{sheet_name}!{coord}"] = "attachment_default"
                        match_summary["alias_matched"] += 1
                        continue

                    financial_value = _match_financial_record(context_labels, section, db)
                    if financial_value:
                        resolved_cells[coord] = financial_value
                        fill_sources[f"{sheet_name}!{coord}"] = "financial_match"
                        match_summary["alias_matched"] += 1
                        continue

                    context_value = _match_company_context_value(context_labels, section, company_context)
                    if context_value:
                        resolved_cells[coord] = context_value
                        fill_sources[f"{sheet_name}!{coord}"] = "context_match"
                        match_summary["alias_matched"] += 1
                        continue

                    best_match = None
                    for label in context_labels:
                        m = matcher.match(label, threshold=88)
                        if m and (not best_match or m.score > best_match.score):
                            best_match = m
                    if best_match:
                        resolved_cells[coord] = best_match.value
                        fill_sources[f"{sheet_name}!{coord}"] = "alias_match"
                        match_summary["alias_matched"] += 1

            protected_set = protected_cells_by_sheet.get(sheet_name, set())
            filtered_cmap_lines = []
            for line in cmap.split('\n'):
                for coord in resolved_cells.keys():
                    line = re.sub(fr'\[{coord}\]=EMPTY', f"[{coord}]='(Already Answered)'", line)
                
                def remove_reserved(m):
                    coord = m.group(1)
                    if not is_allowed_col(coord):
                        return ""
                    if coord in protected_set:
                        return "[" + coord + "]='(Reserved for project/reference picker)'"
                    return m.group(0)
                line = re.sub(r'\[([A-Z]+\d+)\]=EMPTY', remove_reserved, line)
                line = re.sub(r'\s+', ' ', line).strip()
                
                if line:
                    filtered_cmap_lines.append(line)
            filtered_cmap = "\n".join(filtered_cmap_lines)

            if "EMPTY" not in filtered_cmap:
                print(f"[Pass1] Sheet '{sheet_name}' fully resolved without vision call")
                match_summary["sheets_skipped_vision_call"].append(sheet_name)
                all_fills[sheet_name] = resolved_cells
                continue

            b64_images = []
            for path in images_by_sheet.get(sheet_name, []):
                with open(path, "rb") as f:
                    b64_images.append(base64.standard_b64encode(f.read()).decode("utf-8"))
            
            ai_fills = ai_fill_sheet(
                sheet_name, b64_images, filtered_cmap, company_context,
                protected_cells=protected_cells_by_sheet.get(sheet_name, set()),
                reserved_cols=reserved_cols,
                fillable_cols=fillable_cols,
                sheet_form_json=sheet_form_json
            )
            
            for k in ai_fills.keys():
                fill_sources[f"{sheet_name}!{k}"] = "gpt4o_vision"
            match_summary["ai_guessed"] += len(ai_fills)

            all_fills[sheet_name] = {**resolved_cells, **ai_fills}
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

    combined_map = "\n".join(
        f"--- SHEET: {sn} ---\n{sheet_maps[sn]}" for sn in fillable_sheets
    )
    return all_fills, sheet_status, combined_map, match_summary, fill_sources, pending_project_tables


def write_filled_excel_multi(original_bytes: bytes, sheet_fills: dict) -> bytes:
    """
    Write filled values back to their correct sheets, preserving all formatting.
    sheet_fills: {sheet_name: {cell_addr: value}}
    """
    wb = openpyxl.load_workbook(io.BytesIO(original_bytes))
    for sheet_name, fills in sheet_fills.items():
        if sheet_name not in wb.sheetnames:
            print(f"[WARN] Sheet '{sheet_name}' not found in workbook — skipping")
            continue
        ws = wb[sheet_name]
        protected_cells = {c.coordinate for row in ws.iter_rows() for c in row if c.data_type == 'f'}
        
        for addr, value in fills.items():
            if addr in protected_cells:
                print(f"[SKIP] {sheet_name}!{addr} is a formula cell, not overwriting")
                continue
            try:
                cell = ws[addr]
                if type(cell).__name__ == 'MergedCell':
                    anchor_cell = None
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            top_left = merged_range.coord.split(':')[0]
                            anchor_cell = ws[top_left]
                            break
                    anchor_is_blank = anchor_cell is not None and (
                        anchor_cell.value is None or str(anchor_cell.value).strip() == ""
                    )
                    if anchor_is_blank:
                        anchor_cell.value = value
                    else:
                        anchor_desc = f"{anchor_cell.coordinate}='{anchor_cell.value}'" if anchor_cell else "unknown"
                        print(f"[SKIP] {sheet_name}!{addr} is a merged-cell remnant whose anchor "
                              f"already holds label text ({anchor_desc}) -- refusing to overwrite it")
                else:
                    cell.value = value
            except Exception as e:
                print(f"[WARN] Could not write {sheet_name}!{addr}: {e}")
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def write_filled_excel(original_bytes: bytes, sheet_name: str, cell_fills: dict) -> bytes:
    wb = openpyxl.load_workbook(io.BytesIO(original_bytes))
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    for addr, value in cell_fills.items():
        try:
            ws[addr] = value
        except Exception as e:
            print(f"[WARN] Could not write {addr}: {e}")
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _fiscal_year_variants(fy: str) -> list[str]:
    variants = {fy}
    m = re.match(r'^FY\s*(\d{2})-(\d{2})$', fy.strip())
    if m:
        y1, y2 = m.groups()
        variants.add(f"20{y1}-20{y2}")
        variants.add(f"20{y1}-{y2}")
        variants.add(f"FY 20{y1}-{y2}")
        variants.add(f"{y1}-{y2}")
        variants.add(f"FY {y1}-{y2}")
    m2 = re.match(r'^(\d{4})-(\d{4})$', fy.strip())
    if m2:
        y1, y2 = m2.groups()
        variants.add(f"FY{y1[-2:]}-{y2[-2:]}")
        variants.add(f"FY {y1[-2:]}-{y2[-2:]}")
        variants.add(f"{y1[-2:]}-{y2[-2:]}")
    return list(variants)


def build_company_context(db) -> dict:
    from models.database import CompanyField, FinancialRecord, ProjectReference, ProjectFile
    def _clean_dict_value(value):
        if value is None: return ""
        if isinstance(value, str): return value.strip()
        return value

    def _project_ref_payload(ref: ProjectReference) -> dict:
        return {
            "project_name": _clean_dict_value(ref.project_name),
            "client_name": _clean_dict_value(ref.client_name),
            "region": _clean_dict_value(ref.region),
            "location": _clean_dict_value(ref.location),
            "area_sqft": _clean_dict_value(ref.area_sqft),
            "consultant": _clean_dict_value(ref.consultant),
            "pmc": _clean_dict_value(ref.pmc),
            "project_sector": _clean_dict_value(ref.project_sector),
            "project_type": _clean_dict_value(ref.project_type),
            "project_value": _clean_dict_value(ref.project_value),
            "status": _clean_dict_value(ref.status),
            "start_date": _clean_dict_value(ref.start_date),
            "end_date": _clean_dict_value(ref.end_date),
            "client_rep_name": _clean_dict_value(ref.client_rep_name),
            "client_rep_designation": _clean_dict_value(ref.client_rep_designation),
            "client_rep_email": _clean_dict_value(ref.client_rep_email),
            "client_rep_phone": _clean_dict_value(ref.client_rep_phone),
            "certifications": _clean_dict_value(ref.certifications),
        }

    fields = db.query(CompanyField).all()
    financials = db.query(FinancialRecord).all()
    references = db.query(ProjectReference).all()
    documents = db.query(ProjectFile).filter(ProjectFile.source_module == "document").all()

    ctx = {
        "General Information": {},
        "Financial Info": {},
        "Customer References": [],
        "Major Work Done / Projects Done": [],
        "Certificates / Licenses / Attachments": [],
        "Lookup Guidance": "Search the named section first. If the answer is not there, search the full company data across every section before leaving the cell blank.",
    }

    for field in fields:
        if not field.value:
            continue
        section_name = "General Information"
        label = field.field_label or ""
        category = (field.category or "").lower()
        label_lower = label.lower()
        if "financial" in category or any(token in label_lower for token in ["turnover", "balance sheet", "profit", "loss", "assets", "liability", "ebitda", "income", "revenue"]):
            section_name = "Financial Info"
        elif any(token in label_lower for token in ["certificate", "license", "licence", "attachment", "document", "upload"]):
            section_name = "Certificates / Licenses / Attachments"
        if isinstance(ctx[section_name], list):
            ctx[section_name].append({
                "label": label,
                "value": field.value,
                "document_link": field.document_link or "",
                "category": field.category or "",
            })
        else:
            ctx[section_name][label] = field.value

    if financials:
        for f in financials:
            if not f.value:
                continue
            unit_str = f" {f.unit}" if f.unit else ""
            for variant_year in _fiscal_year_variants(f.fiscal_year):
                bucket = ctx["Financial Info"].setdefault(variant_year, {})
                if not isinstance(bucket, dict):
                    # FIX (live bug -- crashed every process-excel call): a
                    # CompanyField can carry a plain label that happens to
                    # collide with a fiscal-year variant string (e.g. an
                    # imported "Financials Turnover" table stored a field
                    # literally labeled "2023-2024"), which set
                    # ctx["Financial Info"]["2023-2024"] to a STRING a few
                    # lines above. setdefault() then returns that existing
                    # string instead of a fresh dict, and item-assignment
                    # into a string raises TypeError. Skip only this one
                    # colliding variant -- _fiscal_year_variants() always
                    # produces several variants per fiscal year, so the data
                    # still reaches the GPT prompt through the others.
                    continue
                bucket[f.metric_label] = f"{f.value}{unit_str}"
        ctx["Financials"] = ctx["Financial Info"]

    for ref in references:
        payload = _project_ref_payload(ref)
        contact_ready = any(payload.get(key) for key in ("client_rep_name", "client_rep_designation", "client_rep_email", "client_rep_phone"))
        project_ready = any(payload.get(key) for key in ("project_name", "client_name", "location", "area_sqft", "project_value", "start_date", "end_date"))

        if contact_ready:
            ctx["Customer References"].append(payload)
        if project_ready:
            ctx["Major Work Done / Projects Done"].append(payload)

    for doc in documents:
        ctx["Certificates / Licenses / Attachments"].append({
            "name": doc.name,
            "doc_type": doc.doc_type,
            "tags": doc.tags or [],
            "has_file": bool(doc.filename),
            "sharepoint_link": doc.sharepoint_link or "",
        })

    return ctx


def get_doc_checklist(text: str, db) -> list:
    from models.database import ProjectFile
    docs = db.query(ProjectFile).filter(ProjectFile.source_module == "document").all()
    # FIX (P1): removed the hardcoded client-specific LOA names ("Taj Hotel LOA",
    # "UBS LOA", "German Consulate LOA") that used to be blindly attached to every
    # single client's checklist whenever the word "completion" appeared anywhere in
    # the form -- nonsensical for any client other than the ones those LOAs actually
    # belong to. Those now map to the generic, reusable doc types instead. Added an
    # "msme" trigger (there was previously no way to ever surface this even though
    # many vendor forms explicitly ask for it), and a working "bank statement"
    # trigger so that DOC_TYPES's existing "Bank Statement" entry is reachable at
    # all (it previously had no trigger word pointing to it -- dead code).
    trigger_map = {
        "gst": ["GST Registration Certificate"],
        "pan": ["Company PAN Card"],
        "msme": ["MSME Registration Certificate"],
        "udyam": ["MSME Registration Certificate"],
        "incorporation": ["Certificate of Incorporation"],
        "iso": ["ISO 45001 Certificate"],
        "turnover": ["Balance Sheet FY 2023-24","Balance Sheet FY 2022-23","Balance Sheet FY 2021-22"],
        "balance": ["Balance Sheet FY 2023-24","Balance Sheet FY 2022-23"],
        "pf": ["PF Registration Certificate"],
        "esi": ["ESI Registration Certificate"],
        "insurance": ["Insurance Certificate"],
        "solvency": ["Solvency Certificate"],
        "quality": ["ISO 45001 Certificate","Quality Policy Document"],
        "organisation": ["Organisation Chart"],
        "organization": ["Organisation Chart"],
        "equipment": ["Plant & Equipment List"],
        "safety": ["EHS Safety Programme"],
        "ohsas": ["EHS Safety Programme"],
        "income tax": ["ITR Certificate"],
        "bank statement": ["Bank Statement"],
        "banker credit": ["Bank Statement"],
        "completion certificate": ["Project Completion Certificate"],
        "appreciation": ["Client Appreciation Letter"],
        "work order": ["Work Order / LOA"],
        "loa": ["Work Order / LOA"],
    }
    mentioned = {"GST Registration Certificate","Company PAN Card","Certificate of Incorporation"}
    tl = text.lower()
    for trigger, docs_list in trigger_map.items():
        if trigger in tl:
            mentioned.update(docs_list)
    checklist, seen = [], set()
    for doc in docs:
        if doc.name in mentioned and doc.name not in seen:
            seen.add(doc.name)
            checklist.append({
                "id": doc.id, "name": doc.name, "doc_type": doc.doc_type,
                "has_file": bool(doc.filename),
                "sharepoint_link": doc.sharepoint_link,
                "download_url": f"/api/documents/download/{doc.id}" if doc.filename else None
            })
    return checklist


# â”€â”€ Routes â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@router.get("/history")
def get_form_history(db: Session = Depends(get_db)):
    forms = db.query(FilledForm).order_by(FilledForm.created_at.desc()).limit(50).all()
    return {"forms": [{k: v for k, v in f.__dict__.items() if not k.startswith("_")} for f in forms]}


@router.get("/{form_id}")
def get_form(form_id: int, db: Session = Depends(get_db)):
    form = db.query(FilledForm).filter(FilledForm.id == form_id).first()
    if not form: raise HTTPException(404, "Form not found")
    return {k: v for k, v in form.__dict__.items() if not k.startswith("_")}


@router.post("/{form_id}/save-answers")
def save_human_answers(form_id: int, answers: dict, db: Session = Depends(get_db)):
    form = db.query(FilledForm).filter(FilledForm.id == form_id).first()
    if not form: raise HTTPException(404, "Form not found")
    merged = dict(form.filled_data or {})
    merged.update(answers)
    form.filled_data = merged
    form.unknown_fields = [f for f in (form.unknown_fields or []) if f not in answers]
    db.commit()
    return {"updated": len(answers), "remaining_unknown": len(form.unknown_fields)}


def _rebuild_filled_bytes(form) -> bytes:
    """Reconstruct sheet_fills from stored 'Sheet!Cell' keys and write to original file."""
    orig_path = os.path.join(UPLOAD_DIR, form.original_filename)
    if not os.path.exists(orig_path):
        raise HTTPException(404, "Original file not found")
    with open(orig_path, "rb") as f:
        original_bytes = f.read()

    sheet_fills = {}
    for key, data in (form.filled_data or {}).items():
        value = data.get("value") if isinstance(data, dict) else data
        if "!" in key:
            sheet_name, cell = key.split("!", 1)
        else:
            # legacy single-sheet data â€” fall back to first sheet
            wb_tmp = openpyxl.load_workbook(io.BytesIO(original_bytes), read_only=True)
            sheet_name = wb_tmp.sheetnames[0]
            wb_tmp.close()
            cell = key
        if re.match(r'^[A-Z]{1,3}\d+$', cell):
            sheet_fills.setdefault(sheet_name, {})[cell] = value

    return write_filled_excel_multi(original_bytes, sheet_fills)


@router.get("/{form_id}/download")
def download_filled_form(form_id: int, db: Session = Depends(get_db)):
    form = db.query(FilledForm).filter(FilledForm.id == form_id).first()
    if not form: raise HTTPException(404, "Form not found")
    filled_bytes = _rebuild_filled_bytes(form)
    out_path = os.path.join(OUTPUT_DIR, f"filled_{form_id}_{form.original_filename}")
    with open(out_path, "wb") as f:
        f.write(filled_bytes)
    return FileResponse(out_path, filename=f"filled_{form.original_filename}")


@router.get("/{form_id}/preview")
def preview_filled_form(form_id: int, db: Session = Depends(get_db)):
    form = db.query(FilledForm).filter(FilledForm.id == form_id).first()
    if not form: raise HTTPException(404, "Form not found")
    filled_bytes = _rebuild_filled_bytes(form)
    out_path = os.path.join(OUTPUT_DIR, f"filled_{form_id}_{form.original_filename}")
    with open(out_path, "wb") as f:
        f.write(filled_bytes)

    try:
        data = read_excel_preview(out_path)
        return {"form_id": form_id, "name": form.original_filename, "sheets": data}
    except Exception as e:
        raise HTTPException(500, f"Could not generate preview: {str(e)}")









