from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Form
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from models.database import (
    get_db, CompanyField, FilledForm,
    ProjectDataRecord, ProjectReference, FinancialRecord, ProjectFile
)
from routers.forms import (
    ai_fill_workbook, write_filled_excel_multi,
    build_company_context, get_doc_checklist,
    openai_client, VISION_MODEL, UPLOAD_DIR, OUTPUT_DIR,
    excel_to_all_sheet_maps, build_workbook_form_json,
    find_field_value_in_record
)
from typing import Optional
import os, json, base64, re, shutil, io
from datetime import datetime
from services.vector_store import VectorStore
import openpyxl

router = APIRouter()

RAG_FALLBACK_THRESHOLD = 0.75

SCREENSHOT_FILL_PROMPT = """You are filling in a Pre-Qualification portal form for HTL Aircon Pvt Ltd.

Look at this screenshot of an online form. For every visible empty field:
- Understand what it is asking
- Find the matching value from the sectioned company data
- Search the most relevant section first, then search the full company data if the answer is not in that section
- Return a JSON array of {label, value} pairs

Example:
[
  {"label": "Company Name", "value": "HTL Aircon Pvt Ltd"},
  {"label": "GST Number", "value": "27AABCH9057L1Z4"}
]

Skip fields you have no data for.
If a field only asks for attachments, certificates, or licenses and does not name a specific document, use "Attached".
Return ONLY the JSON array, no markdown."""


_ROMAN_HEADING_RE = re.compile(r'^[IVXLCDM]{1,4}\s*[\.\)\-:]')
_STOP_LABELS = {
    "remarks", "remark", "sr no", "sr no.", "s no", "s.no", "s.no.",
    "particulars", "particular", "details", "description", "notes", "note",
}


def _is_junk_label(s: str) -> bool:
    """
    Mirrors field_matcher.extract_candidates_from_map's is_usable_label guard.
    FIX (unknown-fields list polluted with section headers): agent.py's own
    adjacent-label lookup below has no equivalent filter, so a roman-numeral
    section heading like "I . GENERAL INFORMATION" (sitting in an unmerged
    cell, so it never gets wrapped as "--- Section: ... ---") was leaking
    into the "Needs your input" review list as if it were a real unanswered
    question -- consuming a slot in the capped 15-item list with a heading
    nobody can actually answer, instead of a genuine blank field.
    """
    if not s:
        return True
    s = s.strip()
    if s.startswith("--- Section:") and s.endswith("---"):
        return False  # legitimate label, not junk
    if _ROMAN_HEADING_RE.match(s):
        return True
    letters = [c for c in s if c.isalpha()]
    if len(s) > 20 and letters and (sum(1 for c in letters if c.isupper()) / len(letters)) > 0.85:
        return True
    return s.lower() in _STOP_LABELS


def _get_adjacent_label_factory(cell_label_map):
    def get_adjacent_label(coord):
        match = re.match(r'^([A-Z]+)(\d+)$', coord)
        if not match: return coord
        col_str, row_str = match.groups()
        row = int(row_str)
        col_idx = 0
        for char in col_str:
            col_idx = col_idx * 26 + (ord(char) - ord('A') + 1)
        for i in range(1, min(col_idx, 5)):
            left_col_idx = col_idx - i
            left_col_str = ""
            temp = left_col_idx
            while temp > 0:
                temp, remainder = divmod(temp - 1, 26)
                left_col_str = chr(65 + remainder) + left_col_str
            left_coord = f"{left_col_str}{row}"
            if left_coord in cell_label_map:
                return cell_label_map[left_coord]
        return coord
    return get_adjacent_label


@router.post("/process-excel")
def process_excel_form(
    client_name: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Fill a (possibly multi-sheet) Excel form like a human would, sheet by sheet:
    1. Build a cell map PER SHEET
    2. Classify each sheet as FILLABLE (has blank fields to fill) or INFO_ONLY
       (cover letters, declarations, auto-calculated summaries â€” left untouched)
    3. Render only FILLABLE sheets to images (one sheet = one set of images)
    4. Send each FILLABLE sheet's image + its own cell map to GPT-4o, so cell
       coordinates never leak across sheets
    5. Write values back to the exact sheet + cell they belong to
    Works for single-sheet OR multi-sheet (e.g. 13-tab) forms alike.
    """
    from utils import sanitize_filename, enforce_upload_size

    file_bytes = file.file.read()
    enforce_upload_size(len(file_bytes))
    safe_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{sanitize_filename(file.filename)}"
    with open(os.path.join(UPLOAD_DIR, safe_name), "wb") as f:
        f.write(file_bytes)

    company_context = build_company_context(db)
    workbook_form_json = build_workbook_form_json(file_bytes)
    structure_name = f"{safe_name}.form_structure.json"
    structure_path = os.path.join(UPLOAD_DIR, structure_name)
    with open(structure_path, "w", encoding="utf-8") as f:
        json.dump(workbook_form_json, f, indent=2, ensure_ascii=False)

    print(f"[1/2] Building per-sheet cell maps + classifying sheets for: {file.filename}")
    try:
        sheet_fills, sheet_status, combined_cell_map, match_summary, fill_sources, pending_project_tables = ai_fill_workbook(
            file_bytes, file.filename, company_context, db, workbook_form_json=workbook_form_json
        )
    except Exception as e:
        raise HTTPException(500, f"Could not process Excel: {str(e)}")

    total_filled_cells = sum(len(f) for f in sheet_fills.values())
    if total_filled_cells == 0 and not pending_project_tables:
        raise HTTPException(
            400,
            "GPT-4o could not fill any fields across any sheet. "
            "Check company data is seeded and the form has recognizable fillable fields."
        )
    print(f"[2/2] Filled {total_filled_cells} cells across {len(sheet_fills)} sheet(s)")

    # Build human-readable display + figure out unknown fields, sheet by sheet
    filled_data = {}
    unknown_fields_objs = []
    total_empty_cells = 0

    for sheet_name, cmap_section in [
        (sn, sec) for sn, sec in (
            (s.split(" ---\n", 1)[0], s)
            for s in combined_cell_map.split("--- SHEET: ")[1:]
        )
    ] if combined_cell_map else []:
        cmap_text = cmap_section.split(" ---\n", 1)[1] if " ---\n" in cmap_section else ""
        cell_label_map = {}
        for line in cmap_text.split("\n"):
            matches = re.findall(r"\[([A-Z]+\d+)\]='([^']*)'", line)
            for coord, val in matches:
                cell_label_map[coord] = val
        get_adjacent_label = _get_adjacent_label_factory(cell_label_map)

        fills = sheet_fills.get(sheet_name, {})
        for cell, value in fills.items():
            label = get_adjacent_label(cell)
            filled_data[f"{sheet_name}!{cell}"] = {
                # FIX (live bug -- garbled labels in the UI): this used to be a
                # literal "—" em-dash, but the source file had it saved with
                # broken encoding (mojibake bytes decoding to "â€”"), so every
                # label shown in Review/Output rendered as garbage. Plain ASCII
                # sidesteps the encoding issue entirely instead of just
                # swapping in a different Unicode character that could suffer
                # the same fate.
                "label": f"{sheet_name} - {label}" if label else f"Cell {cell}",
                "value": value
            }

        all_empty = re.findall(r'\[([A-Z]+\d+)\]=EMPTY', cmap_text)
        total_empty_cells += len(all_empty)
        raw_unknown = [c for c in all_empty if c not in fills]
        for c in raw_unknown:
            lbl = get_adjacent_label(c)
            if lbl and _is_junk_label(lbl):
                continue
            tagged = f"{sheet_name} - {lbl}" if lbl else f"Cell {c}"
            if lbl and lbl != c and tagged not in [u["label"] for u in unknown_fields_objs] and lbl not in [u["label"].split(" - ")[-1] for u in unknown_fields_objs]:
                unknown_fields_objs.append({"label": tagged, "cell": f"{sheet_name}!{c}"})

    unknown_fields_objs = unknown_fields_objs[:15]
    
    # PASS 3: RAG Fallback
    vs = VectorStore()
    enriched_unknown = []
    for field_obj in unknown_fields_objs:
        # field string is usually "SheetName - FieldLabel"
        field = field_obj["label"]
        label_only = field.split(" - ")[-1] if " - " in field else field
        
        results = vs.search(db, query=label_only, top_k=3)
        suggestion = None
        source = None
        if results and results[0]["score"] > RAG_FALLBACK_THRESHOLD:
            top_match = results[0]
            suggestion = top_match["text"]
            # To get source name, we need to query the original file based on source_type and source_id.
            # But just keeping it simple and adding the ID or querying later. Let's query it.
            from models.database import ProjectFile, ProjectReference
            if top_match["source_type"] == "project_file":
                pf = db.query(ProjectFile).filter(ProjectFile.id == top_match["source_id"]).first()
                if pf: source = f"{pf.filename or pf.name}, {top_match['sheet_or_page']}"
            elif top_match["source_type"] == "document":
                doc = db.query(ProjectFile).filter(ProjectFile.id == top_match["source_id"], ProjectFile.source_module == "document").first()
                if doc: source = f"{doc.filename or doc.name}, {top_match['sheet_or_page']}"
            elif top_match["source_type"] == "reference":
                ref = db.query(ProjectReference).filter(ProjectReference.id == top_match["source_id"]).first()
                if ref: source = f"Project Reference: {ref.project_name}"
        
        enriched_unknown.append({
            "label": field,
            "suggested_answer": suggestion,
            "suggested_source": source,
            "cell": field_obj["cell"]
        })

    doc_checklist = get_doc_checklist(combined_cell_map, db)

    stored_fills = {}
    for sheet_name, fills in sheet_fills.items():
        for cell, value in fills.items():
            stored_fills[f"{sheet_name}!{cell}"] = filled_data[f"{sheet_name}!{cell}"]

    sheets_summary = {
        sn: ("Filled" if status == "FILLABLE" and sheet_fills.get(sn) else
             "Skipped (info only)" if status == "INFO_ONLY" else
             "No fields matched")
        for sn, status in sheet_status.items()
    }

    form = FilledForm(
        client_name=client_name,
        form_type="excel",
        original_filename=safe_name,
        filled_data=stored_fills,
        unknown_fields=enriched_unknown,
        doc_checklist=doc_checklist,
        fill_sources=fill_sources,
        pending_project_tables=pending_project_tables
    )
    db.add(form); db.commit(); db.refresh(form)

    return {
        "form_id": form.id,
        "client_name": client_name,
        "total_fields": total_empty_cells,
        "auto_filled": total_filled_cells,
        "unknown_count": len(enriched_unknown),
        "filled_data": filled_data,
        "unknown_fields": enriched_unknown,
        "doc_checklist": doc_checklist,
        "sheets_processed": sheets_summary,
        "match_summary": match_summary,
        "ai_used": "OpenPyXL section JSON + Deterministic Fuzzy/Financial/Attachment Match + Deterministic Vertical-Table Detection + RAG Fallback + GPT-4o vision",
        "pending_project_tables": pending_project_tables,
        "form_structure_json": structure_name
    }


@router.post("/process-image")
def process_image_form(
    client_name: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Screenshot of portal â†’ GPT-4o reads it visually â†’ returns label:value pairs."""
    from utils import enforce_upload_size

    file_bytes = file.file.read()
    enforce_upload_size(len(file_bytes))
    content_type = file.content_type or "image/jpeg"
    company_context = build_company_context(db)
    b64 = base64.standard_b64encode(file_bytes).decode("utf-8")

    # FIX (P0 -- bare 500 with no message on any vision failure): unlike the
    # Excel path (ai_fill_workbook / analyze_sheet), which already tolerates
    # per-sheet vision errors and falls back to Pass 1 matching, this call had
    # zero error handling -- a rate limit, quota exhaustion, timeout, or any
    # other provider hiccup propagated straight into FastAPI's generic 500,
    # showing the user "Request failed with status code 500" with nothing
    # actionable. Surface a clear, specific message instead.
    try:
        response = openai_client.chat.completions.create(
            model=VISION_MODEL,
            max_tokens=4000,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image_url", "image_url": {
                        "url": f"data:{content_type};base64,{b64}",
                        "detail": "high"
                    }},
                    {"type": "text", "text": (
                        f"{SCREENSHOT_FILL_PROMPT}\n\n"
                        f"COMPANY DATA:\n{json.dumps(company_context, indent=2)}"
                    )}
                ]
            }]
        )
    except Exception as e:
        msg = str(e)
        if "429" in msg or "RESOURCE_EXHAUSTED" in msg or "quota" in msg.lower():
            raise HTTPException(
                status_code=503,
                detail="The AI vision service has hit its request quota for now. Please try again in a few minutes, or use the Excel Form upload instead."
            )
        raise HTTPException(
            status_code=502,
            detail=f"The AI vision service failed to process this screenshot: {msg[:200]}"
        )

    raw = response.choices[0].message.content.strip()
    if raw.startswith("```"):
        raw = raw.split("\n", 1)[1].rsplit("```", 1)[0].strip()

    try:
        pairs = json.loads(raw)
        filled_data = {p["label"]: p["value"] for p in pairs if p.get("label") and p.get("value")}
    except Exception:
        filled_data = {}

    doc_checklist = get_doc_checklist(" ".join(filled_data.keys()), db)

    form = FilledForm(
        client_name=client_name,
        form_type="image",
        original_filename=file.filename,
        filled_data=filled_data,
        unknown_fields=[],
        doc_checklist=doc_checklist
    )
    db.add(form); db.commit(); db.refresh(form)

    return {
        "form_id": form.id,
        "client_name": client_name,
        "total_fields": len(filled_data),
        "auto_filled": len(filled_data),
        "unknown_count": 0,
        "filled_data": filled_data,
        "unknown_fields": [],
        "doc_checklist": doc_checklist,
        "ai_used": "GPT-4o vision (GitHub Models)"
    }


@router.post("/save-learned-answer")
def save_learned_answer(
    field_label: str = Form(...),
    answer: str = Form(...),
    form_id: Optional[int] = Form(None),
    save_to_db: bool = Form(True), # Kept for API compatibility
    db: Session = Depends(get_db)
):
    from models.database import CompanyField
    from utils import normalize_field_key
    
    key = normalize_field_key(field_label)
    existing = db.query(CompanyField).filter(CompanyField.field_key == key).first()
    
    if existing:
        existing.value = answer
        existing.confidence = "learned"
        existing.usage_count = (existing.usage_count or 0) + 1
        existing.last_updated = datetime.utcnow()
    else:
        db.add(CompanyField(
            category="Learned", 
            field_key=key,
            field_label=field_label, 
            value=answer,
            confidence="learned",
            usage_count=1,
            aliases=[]
        ))
    db.commit()

    if form_id:
        form = db.query(FilledForm).filter(FilledForm.id == form_id).first()
        if form:
            merged = dict(form.filled_data or {})
            merged[field_label] = {"label": "Manual Answer", "value": answer}
            form.filled_data = merged
            # Remove from unknown fields if it exists (check cell or label).
            # FIX: unknown_fields entries store a display label in "label"
            # (e.g. "Sheet1 — Company Name") and a coordinate in "cell"
            # (e.g. "Sheet1!D6") -- field_label is always the display label, so
            # comparing it against "cell" could never match and this filter
            # never actually removed anything.
            form.unknown_fields = [f for f in (form.unknown_fields or []) if (f.get("label") if isinstance(f, dict) else f) != field_label]
            db.commit()

    return {"saved": True, "field": field_label}


from pydantic import BaseModel
from typing import List

class FillProjectTableRequest(BaseModel):
    sheet_name: str
    table_type: str
    selected_ids: List[int]
    subheading: Optional[str] = None

@router.post("/forms/{form_id}/fill-project-table")
def fill_project_table(
    form_id: int,
    req: FillProjectTableRequest,
    db: Session = Depends(get_db)
):
    form = db.query(FilledForm).filter(FilledForm.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")

    pending_tables = form.pending_project_tables or []

    # FIX (P0 -- wrote to the wrong table): a single sheet very commonly has
    # BOTH a "Client References" table and a "Major Work Done" table pending
    # at once. This used to match on sheet_name alone, so it always grabbed
    # whichever pending table for that sheet came first in the list --
    # regardless of which one the request (and the user, picking from the
    # UI) actually asked to fill. The selected records' data would then get
    # written using the WRONG table's row/column layout, silently landing in
    # some other table's cells while the table the user actually meant to
    # fill was popped off the pending list as if it had been handled, or (if
    # it happened to not be first) stayed marked pending forever with no
    # cells ever written. Match on table_type too (and subheading, when the
    # frontend sends it) so this can't cross-wire two tables on one sheet.
    target_table = None
    target_idx = -1
    for i, pt in enumerate(pending_tables):
        if pt.get("sheet_name") != req.sheet_name or pt.get("table_type") != req.table_type:
            continue
        if req.subheading and pt.get("subheading") != req.subheading:
            continue
        target_table = pt
        target_idx = i
        break

    if not target_table:
        raise HTTPException(status_code=404, detail=f"No pending '{req.table_type}' table found for sheet '{req.sheet_name}'")

    table_fills = {}

    # Real display labels for every known project-data column (e.g. "NAME OF
    # CLIENT", "Contract duration (months)") -- used by find_field_value_in_record
    # to fuzzy-match a canonical field like "client_name" against whatever the
    # source spreadsheet actually called that column, instead of requiring an
    # exact snake_case substring hit. Cheap enough to just always fetch (one
    # query for the whole table, not per record).
    from models.database import ProjectDataColumn
    column_labels = {c.column_key: c.display_label for c in db.query(ProjectDataColumn).all()}

    # FIX (P0/P1): branch for the new deterministic "vertical block" layout (each
    # record's fields are stacked across rows in a single answer column) alongside
    # the pre-existing "horizontal" layout (one column per field, one row per
    # record). See detect_vertical_repeating_blocks() in routers/forms.py.
    if target_table.get("layout") == "vertical":
        answer_col = target_table["answer_column"]
        block_start_rows = target_table["block_start_rows"]
        field_row_offsets = target_table["field_row_offsets"]

        rows_to_fill = min(len(req.selected_ids), len(block_start_rows))
        ids_to_fill = req.selected_ids[:rows_to_fill]

        if req.table_type == "project_reference":
            from models.database import ProjectReference
            refs = db.query(ProjectReference).filter(ProjectReference.id.in_(ids_to_fill)).all()
            ref_dict = {r.id: r for r in refs}
            ordered_refs = [ref_dict[i] for i in ids_to_fill if i in ref_dict]

            field_value_map = {
                "client_name": lambda r: r.client_name,
                "project_name": lambda r: r.project_name,
                "location": lambda r: r.location,
                "area_sqft": lambda r: r.area_sqft,
                "amount": lambda r: r.project_value,
                "start_date": lambda r: r.start_date,
                "completion_date": lambda r: r.end_date,
                "contact_name": lambda r: r.client_rep_name,
                "contact_designation": lambda r: r.client_rep_designation,
                "contact_phone": lambda r: r.client_rep_phone,
                "contact_email": lambda r: r.client_rep_email,
            }
            for ref, block_start in zip(ordered_refs, block_start_rows):
                for field_key, offset in field_row_offsets.items():
                    getter = field_value_map.get(field_key)
                    if getter:
                        val = getter(ref)
                        if val:
                            row_num = block_start + offset
                            table_fills[f"{req.sheet_name}!{answer_col}{row_num}"] = val

        elif req.table_type == "project_details":
            from models.database import ProjectDataRecord
            records = db.query(ProjectDataRecord).filter(ProjectDataRecord.id.in_(ids_to_fill)).all()
            rec_dict = {r.id: r for r in records}
            ordered_records = [rec_dict[i] for i in ids_to_fill if i in rec_dict]

            for rec, block_start in zip(ordered_records, block_start_rows):
                for field_key, offset in field_row_offsets.items():
                    val = find_field_value_in_record(rec.data, field_key, column_labels)
                    if val:
                        row_num = block_start + offset
                        table_fills[f"{req.sheet_name}!{answer_col}{row_num}"] = val

        rows_available = len(block_start_rows)

    else:
        # --- Original "horizontal" layout: one column per field, one row per record ---
        start_row = target_table.get("start_row", 1)
        mapping = target_table.get("mapping", {})

        # Need to re-read the sheet's cell map to get fresh available_rows
        # (since the original pass didn't fill anything here)
        orig_path = os.path.join(UPLOAD_DIR, form.original_filename)
        if not os.path.exists(orig_path):
            raise HTTPException(404, "Original file not found")

        with open(orig_path, "rb") as f:
            file_bytes = f.read()

        sheet_maps, _, _, _ = excel_to_all_sheet_maps(file_bytes)
        cmap = sheet_maps.get(req.sheet_name, "")

        available_rows = []
        first_col = next(iter(mapping.values())) if mapping else None
        if first_col:
            r = start_row
            while f"[{first_col}{r}]=EMPTY" in cmap:
                available_rows.append(r)
                r += 1

        if not available_rows:
            available_rows = list(range(start_row, start_row + len(req.selected_ids)))

        # Cap selection to available rows
        rows_to_fill = min(len(req.selected_ids), len(available_rows))
        ids_to_fill = req.selected_ids[:rows_to_fill]

        if req.table_type == "project_reference":
            from models.database import ProjectReference
            refs = db.query(ProjectReference).filter(ProjectReference.id.in_(ids_to_fill)).all()
            # Preserve selection order
            ref_dict = {r.id: r for r in refs}
            ordered_refs = [ref_dict[i] for i in ids_to_fill if i in ref_dict]

            for ref, row_num in zip(ordered_refs, available_rows):
                if "project_name" in mapping and ref.project_name: table_fills[f"{req.sheet_name}!{mapping['project_name']}{row_num}"] = ref.project_name
                if "client_name" in mapping and ref.client_name: table_fills[f"{req.sheet_name}!{mapping['client_name']}{row_num}"] = ref.client_name
                if "location" in mapping and ref.location: table_fills[f"{req.sheet_name}!{mapping['location']}{row_num}"] = ref.location
                if "area_sqft" in mapping and ref.area_sqft: table_fills[f"{req.sheet_name}!{mapping['area_sqft']}{row_num}"] = ref.area_sqft
                if "amount" in mapping and ref.project_value: table_fills[f"{req.sheet_name}!{mapping['amount']}{row_num}"] = ref.project_value
                if "start_date" in mapping and ref.start_date: table_fills[f"{req.sheet_name}!{mapping['start_date']}{row_num}"] = ref.start_date
                if "completion_date" in mapping and ref.end_date: table_fills[f"{req.sheet_name}!{mapping['completion_date']}{row_num}"] = ref.end_date
                if "contact_name" in mapping and ref.client_rep_name: table_fills[f"{req.sheet_name}!{mapping['contact_name']}{row_num}"] = ref.client_rep_name
                if "contact_designation" in mapping and ref.client_rep_designation: table_fills[f"{req.sheet_name}!{mapping['contact_designation']}{row_num}"] = ref.client_rep_designation
                if "contact_phone" in mapping and ref.client_rep_phone: table_fills[f"{req.sheet_name}!{mapping['contact_phone']}{row_num}"] = ref.client_rep_phone
                if "contact_email" in mapping and ref.client_rep_email: table_fills[f"{req.sheet_name}!{mapping['contact_email']}{row_num}"] = ref.client_rep_email

        elif req.table_type == "project_details":
            from models.database import ProjectDataRecord
            records = db.query(ProjectDataRecord).filter(ProjectDataRecord.id.in_(ids_to_fill)).all()
            rec_dict = {r.id: r for r in records}
            ordered_records = [rec_dict[i] for i in ids_to_fill if i in rec_dict]

            for rec, row_num in zip(ordered_records, available_rows):
                for m_key, m_col in mapping.items():
                    val = find_field_value_in_record(rec.data, m_key, column_labels)
                    if val:
                        table_fills[f"{req.sheet_name}!{m_col}{row_num}"] = val

        rows_available = len(available_rows)

    # Merge into filled_data
    filled_data = dict(form.filled_data or {})
    filled_data.update(table_fills)
    form.filled_data = filled_data
    
    # Remove from pending_tables
    pending_tables.pop(target_idx)
    # Re-assign to trigger SQLAlchemy JSON mutation detection
    form.pending_project_tables = list(pending_tables)
    
    db.commit()
    
    return {
        "filled_cells": len(table_fills),
        "remaining_pending_tables": pending_tables,
        "rows_filled": min(len(req.selected_ids), rows_available),
        "rows_available": rows_available
    }


# ── RFP Questionnaire mode ───────────────────────────────────────────────────
# A third Fill Form mode, distinct from the Excel cell-map and screenshot
# paths above. This one is for a different, narrative-style pre-qual layout:
# one sheet of {ID, Category, Evaluation Criterion, Evidence / Acceptance
# Basis, Weight, VENDOR RESPONSE} rows, grouped into named sections (e.g.
# "Technical Experience & Capability", "Stakeholder & Client References").
# Per the actual database audit done for this exact file (ANJ x GC RFP
# Siemens-Eon-Pune), most of these sections need a TABLE of real records as
# supporting evidence (project history, client references, financials), not
# a single cell value -- so each section gets its own generated sheet, and
# the VENDOR RESPONSE column gets a short, honestly-grounded narrative that
# says exactly how many real records back it up (or says plainly that the
# data doesn't exist yet, for the confirmed gaps -- Pune headcount, ISO
# 9001, VRF OEM, after-sales workshop -- rather than guessing).
_INVALID_SHEET_CHARS = re.compile(r'[\[\]:*?/\\]')


def _safe_sheet_name(name: str) -> str:
    cleaned = _INVALID_SHEET_CHARS.sub('-', str(name or 'Section')).strip()
    return (cleaned or 'Section')[:31]


def _large_projects(db: Session):
    records = db.query(ProjectDataRecord).all()
    rows = []
    for r in records:
        d = r.data or {}
        area = None
        for k in ("project_area_in_sqft", "location_area_sq_ft", "area_sq_ft", "area_in_sqft"):
            v = d.get(k)
            if v:
                try:
                    area = float(str(v).replace(",", ""))
                    break
                except ValueError:
                    pass
        # Upper bound excludes obvious source-data entry errors (a few rows
        # carry area values in the tens of millions of sqft -- no single MEP
        # project is plausibly that size; the largest genuine entries here
        # are ~10 lakh sqft). Without this, junk rows sort to the top of
        # "biggest projects" and crowd out real, verifiable evidence.
        if area and 200000 <= area <= 2000000:
            rows.append({
                "Project Name": d.get("project_name") or d.get("end_user") or r.primary_label,
                "Client / End User": d.get("end_user") or d.get("client_name"),
                "Area (Sq Ft)": area,
                "Contract Value": d.get("contract_value_at_start") or d.get("wo_value_excl_gst"),
                "Completion Date": d.get("end_date"),
            })
    rows.sort(key=lambda x: x["Area (Sq Ft)"] or 0, reverse=True)
    cols = ["Project Name", "Client / End User", "Area (Sq Ft)", "Contract Value", "Completion Date"]
    return cols, rows[:25]


def _lab_rnd_projects(db: Session):
    records = db.query(ProjectDataRecord).filter(ProjectDataRecord.search_text.contains("lab")).limit(200).all()
    rows = []
    for r in records:
        d = r.data or {}
        name = d.get("project_name") or d.get("end_user") or r.primary_label
        if not name:
            continue
        rows.append({
            "Project Name": name,
            "Client / End User": d.get("end_user") or d.get("client_name"),
            "Sector": d.get("job_sector"),
            "Completion Date": d.get("end_date"),
        })
    cols = ["Project Name", "Client / End User", "Sector", "Completion Date"]
    return cols, rows[:25]


def _ongoing_projects(db: Session):
    records = db.query(ProjectDataRecord).filter(ProjectDataRecord.source_file.contains("Ongoing")).limit(25).all()
    rows = []
    for r in records:
        d = r.data or {}
        if not (d.get("project_name") or d.get("client_name")):
            continue
        rows.append({
            "Project Name": d.get("project_name") or d.get("client_name"),
            "Start Date": d.get("start_date"),
            "Expected Completion": d.get("end_date"),
        })
    return ["Project Name", "Start Date", "Expected Completion"], rows


def _pmc_siemens_projects(db: Session):
    records = db.query(ProjectDataRecord).filter(ProjectDataRecord.search_text.contains("siemens")).limit(25).all()
    rows = []
    for r in records:
        d = r.data or {}
        rows.append({
            "Project Name": d.get("project_name") or d.get("client_name"),
            "Client / End User": d.get("end_user") or d.get("client_name"),
            "Location": d.get("location"),
            "Value": d.get("contract_value_at_start") or d.get("wo_value_excl_gst"),
        })
    return ["Project Name", "Client / End User", "Location", "Value"], rows


def _ehs_safety_fields(db: Session):
    fields = db.query(CompanyField).filter(CompanyField.category.in_(["Compliance", "Manpower"])).all()
    rows = [{"Field": f.field_label, "Value": f.value} for f in fields if f.value]
    return ["Field", "Value"], rows


def _financial_history(db: Session, metric_keys=None):
    q = db.query(FinancialRecord)
    if metric_keys:
        q = q.filter(FinancialRecord.metric_key.in_(metric_keys))
    else:
        q = q.filter(FinancialRecord.metric_key.in_(["annual_turnover", "net_worth", "net_profit_after_tax", "total_assets"]))
    records = q.order_by(FinancialRecord.fiscal_year.desc()).all()
    rows = [{"Fiscal Year": r.fiscal_year, "Metric": r.metric_label, "Value": r.value, "Unit": r.unit} for r in records]
    return ["Fiscal Year", "Metric", "Value", "Unit"], rows


def _statutory_documents(db: Session):
    docs = db.query(ProjectFile).filter(ProjectFile.source_module == "document").all()
    rows = [{"Document": d.name, "Category": d.category, "Link": d.sharepoint_link or "Not yet linked"} for d in docs]
    return ["Document", "Category", "Link"], rows


def _client_references(db: Session):
    refs = db.query(ProjectReference).filter(
        ProjectReference.client_rep_name.isnot(None),
        ProjectReference.client_rep_phone.isnot(None),
    ).limit(25).all()
    rows = [{
        "Client Name": r.client_name, "Project": r.project_name,
        "Contact Name": r.client_rep_name, "Designation": r.client_rep_designation,
        "Phone": r.client_rep_phone, "Email": r.client_rep_email,
    } for r in refs]
    return ["Client Name", "Project", "Contact Name", "Designation", "Phone", "Email"], rows


# FIX (P0 -- overclaiming support for questions the data doesn't actually
# answer): a real MEP pre-qual RFP routinely bundles several UNRELATED asks
# under one Category -- e.g. "Execution Planning & Resources" covers ongoing
# projects (data exists), a Pune-specific headcount minimum (doesn't exist
# anywhere in this database), AND an after-sales workshop requirement
# (also doesn't exist). Dispatching purely on the row's CATEGORY, as an
# earlier version of this did, wrote "Supported by 24 records..." into the
# Pune-headcount row too -- true for a neighboring question, false for that
# one. Every question now gets its own evidence lookup based on its actual
# criterion/evidence text, checked in order from most specific to most
# generic, so a category that mixes "have it" and "don't have it" questions
# reports each one honestly instead of borrowing a neighbor's evidence.
def _gather_evidence_for_question(category: str, criterion: str, evidence_text: str, db: Session):
    text = f"{criterion or ''} {evidence_text or ''}".lower()
    cat = (category or "").lower()

    # -- Explicit, confirmed gaps: check these FIRST so they never fall
    # through to a same-category generic match that would wrongly claim
    # support (e.g. "Pune" + "Execution Planning" would otherwise match the
    # ongoing-projects branch below). --
    if "vrf" in text or (" oem" in text or text.startswith("oem")):
        return ("VRF OEM Association", [], [])
    if "pune" in text and any(k in text for k in ("personnel", "payroll", "staff", "employee")):
        return ("Pune Office Staffing", [], [])
    if any(k in text for k in ("after-sales", "after sales", "workshop")):
        return ("After-Sales Service Team", [], [])
    if "9001" in text or ("qms" in text and "iso" in text):
        return ("QMS ISO 9001 Certification", [], [])
    if "separate team" in text and any(k in text for k in ("testing", "commissioning", "handover")):
        # The generic EHS/Quality fields (ISO 45001, safety/quality
        # headcounts) don't confirm a DEDICATED T&C/handover team exists --
        # that's a distinct organisational fact this database doesn't track.
        return ("Dedicated Testing & Commissioning Team", [], [])

    # -- Specific evidence types that DO exist -- checked before the
    # generic per-category fallback so they don't get merged together. --
    if "siemens" in text or ("pmc" in text and ("reputed" in text or "experience" in text)):
        cols, rows = _pmc_siemens_projects(db)
        return ("PMC & Siemens Client Experience", cols, rows)
    if any(k in text for k in ("laboratory", " lab ", "lab/", "r&d", "critical environment")):
        cols, rows = _lab_rnd_projects(db)
        return ("Lab/R&D Project History", cols, rows)
    if any(k in text for k in ("ongoing project", "workload", "current project", "current deployment")):
        cols, rows = _ongoing_projects(db)
        return ("Ongoing Projects", cols, rows)
    if "turnover" in text:
        cols, rows = _financial_history(db, metric_keys=["annual_turnover"])
        return ("Turnover History", cols, rows)

    # -- Generic per-category fallback for questions that don't hit any of
    # the specific patterns above. --
    if "technical" in cat or "experience" in cat or "capability" in cat:
        cols, rows = _large_projects(db)
        return ("Large-Scale Project History", cols, rows)
    if "execution" in cat or "resource" in cat or "planning" in cat:
        cols, rows = _ongoing_projects(db)
        return ("Ongoing Projects", cols, rows)
    if "ehs" in cat or "safety" in cat or "quality" in cat or "testing" in cat or "handover" in cat:
        cols, rows = _ehs_safety_fields(db)
        return ("EHS & Quality Records", cols, rows)
    if "financial" in cat:
        cols, rows = _financial_history(db)
        return ("Financial Statements", cols, rows)
    if "governance" in cat or "documentation" in cat:
        cols, rows = _statutory_documents(db)
        return ("Statutory Documents", cols, rows)
    if "stakeholder" in cat or "reference" in cat or "client" in cat:
        cols, rows = _client_references(db)
        return ("Client References", cols, rows)

    return (None, [], [])


def _build_narrative(rows: list, sheet_name: str | None) -> str:
    if not rows:
        return ("No supporting data currently exists in the company database for this requirement. "
                "This needs to be supplied manually before submission.")
    return (f"Supported by {len(rows)} record(s) from HTL's project/company database. "
            f"See the '{sheet_name}' sheet in this workbook for the full supporting detail.")


@router.post("/process-rfp-questionnaire")
def process_rfp_questionnaire(
    client_name: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    from utils import sanitize_filename, enforce_upload_size

    file_bytes = file.file.read()
    enforce_upload_size(len(file_bytes))
    safe_name = sanitize_filename(file.filename)

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.worksheets[0]

    # Locate the header row by finding "VENDOR RESPONSE" (or similar) --
    # scans the first 10 rows since these RFP templates usually have a
    # title/instructions banner above the real header.
    header_row = None
    col_map = {}
    for row in ws.iter_rows(min_row=1, max_row=10):
        found = {}
        for cell in row:
            if not cell.value:
                continue
            h = str(cell.value).upper()
            if "RESPONSE" in h: found["response"] = cell.column
            elif "CATEGORY" in h: found["category"] = cell.column
            elif "CRITERION" in h: found["criterion"] = cell.column
            elif "EVIDENCE" in h: found["evidence"] = cell.column
            elif "WEIGHT" in h: found["weight"] = cell.column
            elif h.strip() == "ID": found["id"] = cell.column
        # Require "response" together with at least one other structural
        # column on the SAME row before accepting it as the header -- a
        # plain "RESPONSE" substring match alone also fires on instructional
        # banner text above the real header (e.g. "...WE REQUIRE A WRITTEN
        # RESPONSE..."), which isn't itself a header row.
        if "response" in found and ("category" in found or "criterion" in found or "evidence" in found):
            header_row = row[0].row
            col_map = found
            break

    if header_row is None:
        raise HTTPException(400, "Could not find a 'VENDOR RESPONSE' column in this sheet -- this endpoint expects the ID / Category / Evaluation Criterion / Evidence / Weight / VENDOR RESPONSE layout.")

    questions = []
    for r in range(header_row + 1, ws.max_row + 1):
        crit_col = col_map.get("criterion") or col_map.get("evidence")
        crit = ws.cell(row=r, column=crit_col).value if crit_col else None
        if not crit:
            continue
        cat_col = col_map.get("category")
        ev_col = col_map.get("evidence")
        questions.append({
            "row": r,
            "category": (ws.cell(row=r, column=cat_col).value if cat_col else None) or "General",
            "criterion": crit,
            "evidence_text": ws.cell(row=r, column=ev_col).value if ev_col else None,
        })

    if not questions:
        raise HTTPException(400, "No question rows found under the header row.")

    # Cache per evidence-TYPE label (not per category) so two questions that
    # both need e.g. "Ongoing Projects" share one sheet instead of each
    # question re-running the query and creating a duplicate.
    evidence_cache = {}
    category_summary = {}
    for q in questions:
        label, columns, rows = _gather_evidence_for_question(q["category"], q["criterion"], q["evidence_text"], db)

        if label not in evidence_cache:
            sheet_name = None
            if rows:
                sheet_name = _safe_sheet_name(label)
                base_name, n = sheet_name, 1
                while sheet_name in wb.sheetnames:
                    n += 1
                    sheet_name = _safe_sheet_name(f"{base_name[:28]}-{n}")
                evsheet = wb.create_sheet(sheet_name)
                for ci, col_name in enumerate(columns, start=1):
                    evsheet.cell(row=1, column=ci, value=col_name)
                for ri, rec in enumerate(rows, start=2):
                    for ci, col_name in enumerate(columns, start=1):
                        evsheet.cell(row=ri, column=ci, value=rec.get(col_name))
            evidence_cache[label] = (sheet_name, rows)

        sheet_name, rows = evidence_cache[label]
        narrative = _build_narrative(rows, sheet_name)
        ws.cell(row=q["row"], column=col_map["response"], value=narrative)

        cat_entry = category_summary.setdefault(q["category"], {"questions": 0, "evidence_types": {}})
        cat_entry["questions"] += 1
        cat_entry["evidence_types"][label or "None"] = {"evidence_rows": len(rows), "sheet_created": sheet_name}

    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_bytes = out_buf.getvalue()

    out_filename = f"rfp_filled_{safe_name}"
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(os.path.join(OUTPUT_DIR, out_filename), "wb") as f:
        f.write(out_bytes)

    form = FilledForm(
        client_name=client_name,
        form_type="rfp_questionnaire",
        original_filename=safe_name,
        filled_data={"category_summary": category_summary, "output_filename": out_filename},
        unknown_fields=[],
        doc_checklist=[],
    )
    db.add(form); db.commit(); db.refresh(form)

    return {
        "form_id": form.id,
        "client_name": client_name,
        "category_summary": category_summary,
        "download_url": f"/api/agent/rfp-download/{form.id}",
    }


@router.get("/rfp-download/{form_id}")
def download_rfp_result(form_id: int, db: Session = Depends(get_db)):
    form = db.query(FilledForm).filter(FilledForm.id == form_id, FilledForm.form_type == "rfp_questionnaire").first()
    if not form:
        raise HTTPException(404, "Form not found")
    out_filename = (form.filled_data or {}).get("output_filename")
    path = os.path.join(OUTPUT_DIR, out_filename) if out_filename else None
    if not path or not os.path.exists(path):
        raise HTTPException(404, "Output file missing")
    return FileResponse(path, filename=out_filename)


