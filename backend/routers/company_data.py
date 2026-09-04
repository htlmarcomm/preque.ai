from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from models.database import get_db, CompanyField
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime
import openpyxl, io, json, os, re
from utils import normalize_field_key

router = APIRouter()

# ── Schemas ──────────────────────────────────────────────────────────────
class FieldCreate(BaseModel):
    category: str
    field_key: str
    field_label: str
    value: Optional[str] = ""
    document_link: Optional[str] = ""
    notes: Optional[str] = ""
    aliases: Optional[List[str]] = []

class FieldUpdate(BaseModel):
    value: Optional[str] = None
    document_link: Optional[str] = None
    notes: Optional[str] = None
    field_label: Optional[str] = None
    category: Optional[str] = None
    aliases: Optional[List[str]] = None

# ── CATEGORIES ────────────────────────────────────────────────────────────
CATEGORIES = [
    "Company Identity", "Address & Contact", "Contact Persons",
    "Registration & Legal", "Manpower", "Infrastructure",
    "Financial", "Business Profile & Capability", "Compliance"
]

@router.get("/categories")
def get_categories():
    return {"categories": CATEGORIES}

# ── ALL FIELDS ────────────────────────────────────────────────────────────
@router.get("/fields")
def get_all_fields(category: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(CompanyField)
    if category:
        q = q.filter(CompanyField.category == category)
    fields = q.order_by(CompanyField.category, CompanyField.id).all()
    return {"fields": [f.__dict__ for f in fields]}

@router.post("/fields")
def create_field(data: FieldCreate, db: Session = Depends(get_db)):
    existing = db.query(CompanyField).filter(CompanyField.field_key == data.field_key).first()
    if existing:
        raise HTTPException(400, f"Field key '{data.field_key}' already exists")
    field = CompanyField(**data.dict())
    db.add(field)
    db.commit()
    db.refresh(field)
    return field.__dict__

@router.put("/fields/{field_id}")
def update_field(field_id: int, data: FieldUpdate, db: Session = Depends(get_db)):
    field = db.query(CompanyField).filter(CompanyField.id == field_id).first()
    if not field:
        raise HTTPException(404, "Field not found")
    for k, v in data.dict(exclude_none=True).items():
        setattr(field, k, v)
    field.last_updated = datetime.utcnow()
    db.commit()
    db.refresh(field)
    return field.__dict__

@router.delete("/fields/{field_id}")
def delete_field(field_id: int, db: Session = Depends(get_db)):
    field = db.query(CompanyField).filter(CompanyField.id == field_id).first()
    if not field:
        raise HTTPException(404, "Field not found")
    db.delete(field)
    db.commit()
    return {"deleted": field_id}

# SECURITY: a `/dump` endpoint that returned every company_fields row (GSTIN,
# PAN, banking limits, financials -- everything) as one unauthenticated-in-
# practice GET used to live here. Nothing in this codebase called it -- not
# the frontend, not build_company_context (which queries the DB directly) --
# it was pure dead code that happened to also be the single highest-value
# target in the app: anyone with the shared API key (baked into the public
# frontend JS bundle, readable via view-source by any visitor) could
# exfiltrate the entire company dataset in one request. Removed rather than
# access-restricted since nothing depends on it.

# ── IMPORT FROM EXCEL ─────────────────────────────────────────────────────
@router.post("/import-excel")
async def import_from_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    import re as _re
    from utils import enforce_upload_size
    contents = await file.read()
    enforce_upload_size(len(contents))
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active
    imported = 0
    skipped = 0
    errors = []

    # Collect all rows to auto-detect whether column A is a serial-number column.
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return {"imported": 0, "updated": 0, "errors": []}

    # Heuristic: if > 60% of non-empty column-A values are purely numeric,
    # treat column A as a serial-number column and shift label/value to B/C.
    col_a_vals = [str(r[0]).strip() for r in all_rows if r and r[0] is not None]
    numeric_count = sum(1 for v in col_a_vals if _re.fullmatch(r'\d+', v))
    has_serial_col = len(col_a_vals) > 0 and (numeric_count / len(col_a_vals)) > 0.6

    # Also skip common header words in the first row
    HEADER_WORDS = {"field", "label", "sr", "sr.", "sr no", "s.no", "s no", "sl",
                    "sl.", "sl no", "no.", "sno", "serial", "particulars", "description",
                    "parameter", "details", "item", ""}

    current_section = "Basic Info"

    for row in all_rows:
        if not row:
            continue

        # Determine label_col and value_col based on serial-number detection
        if has_serial_col:
            if len(row) < 2 or row[1] is None:
                continue
            raw_label = str(row[1]).strip()
            raw_value = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            doc_link = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        else:
            if row[0] is None:
                continue
            raw_label = str(row[0]).strip()
            raw_value = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            doc_link = str(row[2]).strip() if len(row) > 2 and row[2] else ""

        if not raw_label:
            continue

        # Skip header rows
        if raw_label.lower().strip() in HEADER_WORDS:
            continue

        # Detect section-header rows: label present but value is empty and the
        # label looks like a section title (e.g. all-caps, or ends with colon,
        # or is a known category-style phrase).
        if not raw_value:
            label_lower = raw_label.lower()
            looks_like_section = (
                raw_label.isupper() or
                raw_label.endswith(":") or
                any(kw in label_lower for kw in [
                    "general information", "financial", "contact", "address",
                    "registration", "legal", "manpower", "infrastructure",
                    "compliance", "business profile", "capability", "identity",
                    "project", "reference", "certification", "bank",
                ])
            )
            if looks_like_section:
                current_section = raw_label.strip().rstrip(":")
                continue

        field_key = normalize_field_key(raw_label)
        if not field_key or _re.fullmatch(r'\d+', field_key):
            # Pure-numeric keys are serial numbers that slipped through, skip
            continue

        try:
            existing = db.query(CompanyField).filter(CompanyField.field_key == field_key).first()
            if existing:
                existing.value = raw_value
                existing.document_link = doc_link
                existing.last_updated = datetime.utcnow()
                skipped += 1
            else:
                db.add(CompanyField(
                    category=current_section,
                    field_key=field_key,
                    field_label=raw_label,
                    value=raw_value,
                    document_link=doc_link
                ))
                db.flush()  # catch IntegrityError per-row, not at the end
                imported += 1
        except Exception as e:
            db.rollback()
            errors.append(f"Row '{raw_label}': {str(e)[:120]}")
            continue

    db.commit()
    result = {"imported": imported, "updated": skipped}
    if errors:
        result["errors"] = errors[:10]  # return first 10 errors max
    return result

# FIX (P0 -- import silently produces zero rows for common real-world
# headers): this used to match each column header against a single
# hardcoded substring per field ("name of project", "start", "phone"...).
# Real client-supplied templates vary a lot -- "Name of the project" (note
# "the") never matched "name of project", "Date of commencement" never
# matched "start", "Mobile No." never matched "phone"/"contact", and
# "Name and address of the organisation" never matched "client" at all.
# Since project_name is required for a row to be kept at all, a template
# using any of these ordinary phrasings imported ZERO rows with no visible
# error. Each field now has a list of real-world phrasings, matched via the
# same fuzzy scorer (with the same length-aware threshold + ambiguity guard
# to avoid false positives) already proven on the Fill Form project-table
# picker for exactly this class of problem.
PROJECT_REFERENCE_FIELD_KEYWORDS = {
    "project_name": ["project name", "name of project", "name of the project", "project title"],
    "client_name": ["client name", "customer name", "organisation", "organization", "employer", "name and address of the organisation"],
    "region": ["region", "zone"],
    "location": ["location", "town", "locality"],
    "area_sqft": ["area sqft", "area (sqft)", "area"],
    "consultant": ["consultant", "architect"],
    "pmc": ["pmc"],
    "project_sector": ["sector"],
    "project_type": ["project type", "type of project"],
    "project_value": ["project value", "project cost", "contract value", "value of project", "value"],
    "status": ["status"],
    "start_date": ["start date", "date of commencement", "commencement"],
    "end_date": ["end date", "completion date", "date of completion"],
    "client_rep_designation": ["designation"],
    "client_rep_email": ["email"],
    "client_rep_phone": ["phone", "contact", "mobile", "landline"],
    "certifications": ["certification", "certificate"],
}


PLACEHOLDER_VALUES = {"-", "--", "na", "n/a", "nil", "none", "nan", "."}


def _is_placeholder(val_str: str) -> bool:
    """Source spreadsheets commonly fill an unknown cell with a bare '-' or
    'NA' rather than leaving it empty -- storing that literally as e.g. a
    client rep name is worse than just leaving the field blank."""
    return val_str.strip().lower() in PLACEHOLDER_VALUES


def _best_project_reference_field(header_text: str) -> str | None:
    try:
        from rapidfuzz import fuzz
    except ImportError:
        fuzz = None

    def norm(s):
        return re.sub(r'[^a-z0-9]+', ' ', str(s or '').lower()).strip()

    candidate = norm(header_text)
    if not candidate:
        return None

    candidate_tokens = candidate.split()
    required = 92 if len(candidate_tokens) <= 2 else 82

    # (score, keyword length in tokens, field) for every keyword that clears
    # the threshold -- a header can legitimately contain a full match for
    # more than one field's keyword at once (e.g. "Name of the Project
    # (with client name)" scores 100 against BOTH "name of the project" and
    # "client name"), so length is tracked to break that tie below.
    hits = []
    for field, keywords in PROJECT_REFERENCE_FIELD_KEYWORDS.items():
        for kw in keywords:
            kw_norm = norm(kw)
            score = fuzz.token_set_ratio(kw_norm, candidate) if fuzz else (100 if kw_norm in candidate else 0)
            if score >= required:
                hits.append((score, len(kw_norm.split()), field))

    if not hits:
        return None

    # Prefer the highest score; among near-ties, prefer the longer (more
    # specific, less likely to be a coincidental partial mention) keyword.
    hits.sort(key=lambda x: (-x[0], -x[1]))
    best_score, best_len, best_field = hits[0]

    # Ambiguity guard: a DIFFERENT field whose match is both similarly
    # strong AND at least as specific means this is a genuine coin flip --
    # refuse rather than risk misfiling the column.
    for score, kw_len, field in hits[1:]:
        if field != best_field and (best_score - score) < 8 and kw_len >= best_len:
            return None

    return best_field


@router.post("/import-projects")
async def import_projects_from_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    from models.database import ProjectReference
    from utils import enforce_upload_size
    contents = await file.read()
    enforce_upload_size(len(contents))
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    imported = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(h).strip().lower() if h else "" for h in row]
                continue

            if not any(row):
                continue

            row_data = {}
            for h, val in zip(headers, row):
                if not h: continue
                val_str = str(val).strip() if val is not None else ""
                # The combined "client rep" mega-column (name/designation/email/
                # phone all in one cell) is a structural pattern, not a field
                # name -- check it before the generic fuzzy field match.
                # FIX (P0 -- designation overwrote the correctly-parsed name):
                # a header like "Designation of client representative" also
                # contains "representative", so it used to trip this branch
                # too. Its value ("(Procurement Lead)") doesn't look like a
                # combined name/designation/email/phone blob, so it fell into
                # the "else: treat whole value as the rep's name" case and
                # clobbered whatever "Name of the client representative"
                # (processed earlier in the same row) had already set.
                # Requiring "name" alongside "rep"/"representative" limits
                # this branch to genuinely name-shaped columns; a column
                # that's specifically about designation/phone/email now falls
                # through to the ordinary fuzzy field match instead.
                if ("rep" in h or "representative" in h) and "name" in h:
                    # FIX: val_str is already .strip()'d above, so a trailing
                    # "\n" that used to gate this extraction (the old check
                    # was `":" in val_str or "\n" in val_str`) is gone before
                    # it's ever inspected -- "Name - Naveen Sharma\n" reaches
                    # here as "Name - Naveen Sharma" with no newline and no
                    # colon, so the gate always failed and the raw
                    # "Name - Naveen Sharma" got stored verbatim instead of
                    # just "Naveen Sharma". Try the "Name: ..." extraction
                    # directly instead of gating on a separator character
                    # that may or may not have survived stripping.
                    name_m = re.search(r'Name\s*[:\-]?\s*(.*?)(?:\s*Designation|\s*Email|\s*Contact|$)', val_str, re.IGNORECASE | re.DOTALL)
                    if name_m and name_m.group(1).strip():
                        row_data["client_rep_name"] = name_m.group(1).strip()
                        desig_m = re.search(r'Designation\s*[:\-]?\s*(.*?)(?:\s*Email|\s*Contact|$)', val_str, re.IGNORECASE | re.DOTALL)
                        if desig_m and desig_m.group(1).strip(): row_data["client_rep_designation"] = desig_m.group(1).strip()

                        email_m = re.search(r'Email(?: ID)?\s*[:\-]?\s*(.*?)(?:\s*Contact|$)', val_str, re.IGNORECASE | re.DOTALL)
                        if email_m and email_m.group(1).strip(): row_data["client_rep_email"] = email_m.group(1).strip()

                        phone_m = re.search(r'Contact(?: Number)?\s*[:\-]?\s*(.*)', val_str, re.IGNORECASE | re.DOTALL)
                        if phone_m and phone_m.group(1).strip(): row_data["client_rep_phone"] = phone_m.group(1).strip()
                    elif not _is_placeholder(val_str):
                        row_data["client_rep_name"] = val_str
                    continue

                field = _best_project_reference_field(h)
                if field and val_str and not _is_placeholder(val_str):
                    row_data[field] = val_str

            if not row_data.get("project_name"):
                continue
                
            # check if exists
            existing = db.query(ProjectReference).filter(ProjectReference.project_name == row_data["project_name"]).first()
            if existing:
                for k, v in row_data.items():
                    setattr(existing, k, v)
            else:
                row_data["source_file"] = file.filename
                db.add(ProjectReference(**row_data))
            imported += 1
            
    db.commit()
    return {"imported": imported}

# ── SEED FROM HTL DATA ────────────────────────────────────────────────────
@router.post("/seed")
def seed_htl_data(db: Session = Depends(get_db)):
    """Seed DB with deduplicated HTL pre-qual data."""
    existing = db.query(CompanyField).count()
    if existing > 0:
        return {"message": "DB already seeded", "count": existing}

    seed_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "seed_data", "company_seed.json")
    if not os.path.exists(seed_path):
        raise HTTPException(
            500,
            "Seed data file not found at backend/seed_data/company_seed.json. "
            "This file holds company-identifying data (GSTIN, PAN, contact numbers) "
            "and is intentionally gitignored -- ask an admin for a copy."
        )
    with open(seed_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    for item in data:
        db.add(CompanyField(**item))
    db.commit()
    return {"seeded": len(data)}
