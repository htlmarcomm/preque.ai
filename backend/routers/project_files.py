from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Form, BackgroundTasks
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from models.database import get_db, Base, engine, DocumentChunk, ProjectFile
from pydantic import BaseModel
from typing import Optional
import openpyxl, io, os, json, logging
from datetime import datetime
from services.doc_extractor import extract_text, chunk_text
from services.vector_store import VectorStore

router = APIRouter()
FILES_DIR = "uploads/project_files"
os.makedirs(FILES_DIR, exist_ok=True)
logger = logging.getLogger(__name__)

CATEGORIES = [
    "Project Registry", 
    "Client Specific Data", 
    "Company General Data", 
    "Company Financial Data",
    "Company Compliance Data",
    "Employee Details",
    "Company Reports",
    "Project Completion Certificate and Appreciation",
    "Safety and HSE",
    "Policies"
]

class FileUpdate(BaseModel):
    name:             Optional[str] = None
    client:           Optional[str] = None
    category:         Optional[str] = None
    sharepoint_link:  Optional[str] = None
    tags:             Optional[list] = None
    notes:            Optional[str]  = None

def read_excel_preview(filepath: str) -> dict:
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    preview = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        headers = []
        header_set = False
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            cleaned = [str(v).strip() if v is not None else "" for v in row]
            if not any(cleaned):
                continue
            if not header_set:
                headers = cleaned
                header_set = True
            else:
                rows.append(cleaned)
            if i > 5000:
                break
        preview[sheet_name] = {
            "headers": headers,
            "rows": rows[:5000],
            "total_rows": len(rows)
        }
    wb.close()
    return preview

@router.get("/categories")
def get_categories():
    return {"categories": CATEGORIES}

@router.get("/")
def list_files(
    client:   Optional[str] = None,
    category: Optional[str] = None,
    search:   Optional[str] = None,
    db: Session = Depends(get_db)
):
    q = db.query(ProjectFile)
    if client:   q = q.filter(ProjectFile.client.ilike(f"%{client}%"))
    if category: q = q.filter(ProjectFile.category == category)
    if search:
        q = q.filter(
            ProjectFile.name.ilike(f"%{search}%") |
            ProjectFile.client.ilike(f"%{search}%") |
            ProjectFile.notes.ilike(f"%{search}%")
        )
    files = q.order_by(ProjectFile.uploaded_at.desc()).all()
    return {"files": [
        {k: v for k, v in f.__dict__.items() if not k.startswith("_")}
        for f in files
    ]}

@router.post("/upload")
async def upload_file(
    background_tasks: BackgroundTasks,
    name:            str  = Form(...),
    client:          str  = Form(""),
    category:        str  = Form("Company General Data"),
    sharepoint_link: str  = Form(""),
    tags:            str  = Form(""),
    notes:           str  = Form(""),
    file: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db)
):
    filename   = None
    sheet_names = []
    row_count  = 0
    dest = None
    if file:
        safe = file.filename.replace(" ", "_")
        dest = os.path.join(FILES_DIR, safe)
        contents = await file.read()
        with open(dest, "wb") as f:
            f.write(contents)
        filename = safe
        try:
            wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
            sheet_names = wb.sheetnames
            ws = wb.active
            row_count = ws.max_row or 0
            wb.close()
        except Exception:
            pass
    tag_list = [t.strip() for t in tags.split(",") if t.strip()]
    pf = ProjectFile(
        name=name, client=client, category=category,
        filename=filename, sharepoint_link=sharepoint_link,
        tags=tag_list, sheet_names=sheet_names,
        row_count=row_count, notes=notes
    )
    db.add(pf)
    db.commit()
    db.refresh(pf)
    
    if dest:
        try:
            pages = extract_text(dest)
            chunks = chunk_text(pages)
            for chunk in chunks:
                doc_chunk = DocumentChunk(
                    source_type="project_file",
                    source_id=pf.id,
                    sheet_or_page=chunk["sheet_or_page"],
                    chunk_index=chunk["chunk_index"],
                    text=chunk["text"]
                )
                db.add(doc_chunk)
            db.commit()
            background_tasks.add_task(VectorStore().embed_missing, db)
        except Exception as e:
            logger.warning(f"Failed to extract text for project file {pf.id}: {e}")

    return {k: v for k, v in pf.__dict__.items() if not k.startswith("_")}


@router.put("/{file_id}")
def update_file(file_id: int, data: FileUpdate, db: Session = Depends(get_db)):
    pf = db.query(ProjectFile).filter(ProjectFile.id == file_id).first()
    if not pf: raise HTTPException(404, "File not found")
    for k, v in data.dict(exclude_none=True).items():
        setattr(pf, k, v)
    db.commit(); db.refresh(pf)
    return {k: v for k, v in pf.__dict__.items() if not k.startswith("_")}

@router.delete("/{file_id}")
def delete_file(file_id: int, db: Session = Depends(get_db)):
    pf = db.query(ProjectFile).filter(ProjectFile.id == file_id).first()
    if not pf: raise HTTPException(404, "File not found")
    if pf.filename:
        p = os.path.join(FILES_DIR, pf.filename)
        if os.path.exists(p): os.remove(p)
    db.delete(pf); db.commit()
    return {"deleted": file_id}

@router.get("/{file_id}/preview")
def preview_file(file_id: int, db: Session = Depends(get_db)):
    pf = db.query(ProjectFile).filter(ProjectFile.id == file_id).first()
    if not pf: raise HTTPException(404, "File not found")
    if not pf.filename:
        raise HTTPException(400, "No file uploaded — only SharePoint link stored")
    path = os.path.join(FILES_DIR, pf.filename)
    if not os.path.exists(path):
        raise HTTPException(404, "File missing from disk")
    try:
        data = read_excel_preview(path)
        return {"file_id": file_id, "name": pf.name, "sheets": data}
    except Exception as e:
        raise HTTPException(500, f"Could not read Excel: {str(e)}")

@router.get("/{file_id}/download")
def download_file(file_id: int, db: Session = Depends(get_db)):
    pf = db.query(ProjectFile).filter(ProjectFile.id == file_id).first()
    if not pf or not pf.filename: raise HTTPException(404, "File not found")
    path = os.path.join(FILES_DIR, pf.filename)
    if not os.path.exists(path): raise HTTPException(404, "File missing")
    return FileResponse(path, filename=pf.filename)

@router.get("/{file_id}/view")
def view_file_inline(file_id: int, db: Session = Depends(get_db)):
    pf = db.query(ProjectFile).filter(ProjectFile.id == file_id).first()
    if not pf or not pf.filename: raise HTTPException(404, "File not found")
    path = os.path.join(FILES_DIR, pf.filename)
    if not os.path.exists(path): raise HTTPException(404, "File missing")
    return FileResponse(path, content_disposition_type="inline")

@router.post("/add-sharepoint")
def add_sharepoint_link(
    name:            str = Form(...),
    client:          str = Form(""),
    category:        str = Form("Company General Data"),
    sharepoint_link: str = Form(...),
    tags:            str = Form(""),
    notes:           str = Form(""),
    db: Session = Depends(get_db)
):
    tag_list = [t.strip() for t in tags.split(",") if t.strip()]
    pf = ProjectFile(
        name=name, client=client, category=category,
        sharepoint_link=sharepoint_link, tags=tag_list, notes=notes
    )
    db.add(pf); db.commit(); db.refresh(pf)
    return {k: v for k, v in pf.__dict__.items() if not k.startswith("_")}
